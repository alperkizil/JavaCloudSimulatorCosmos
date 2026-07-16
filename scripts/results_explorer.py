#!/usr/bin/env python3
"""
COSMOS Experiment Results Explorer — interactive GUI for campaign result folders.

Manually invoked (never auto-run by the pipeline):

    python scripts/results_explorer.py [results/<ExperimentId>]

Loads a campaign result folder (the directory holding scenario_N_*.csv,
experiment_summary.csv, plot_options.json) and provides:

  * Scatter tab   — universal Pareto front (toggleable, connected), per-
                    algorithm pooled fronts (union over seeds, connected),
                    optional raw per-seed point clouds, X/Y swap, custom title,
                    live per-algorithm colors, family-stable markers,
                    legend / per-algorithm label toggles, box + wheel zoom.
  * Metrics tab   — bar charts for Hypervolume / GD / IGD / Spacing / Pareto
                    contribution, one panel per scenario, with an Aggregation
                    switch: "Collective" (default; indicators of each
                    algorithm's pooled union-over-seeds front) or "Mean ± std
                    over seeds" (per-seed averages).
  * Runtime tab   — average runtime per algorithm, per scenario + overall.
  * Tables tab    — the Global metrics table (default: one row per algorithm,
                    indicators of the collective pooled front) plus per-seed
                    pivot tables (all seeds + MEAN/STDDEV + universal info).
  * Details tab   — per-solution schedule details (task→VM assignments, per-VM
                    queue sizes, per-host energy, makespan / wait / energy
                    aggregates) read from scenario_N_solution_details.json when
                    the run was produced by a build that exports it.
  * Save          — any figure to PNG/SVG/PDF at a chosen DPI, any table to
                    CSV or Excel, every table into one formatted Excel workbook
                    (needs openpyxl), and a "Save for Claude" bundle for LLM
                    analysis: a zip with a generated read-first briefing, every
                    figure, every table as a real CSV, and the all-tables
                    markdown (a plain .md/.md.gz save is still offered).

Simplicity rules (owner decision, 2026-07-16): the GUI shows ONE hypervolume —
the fixed-reference HV_fixed (recompute_hv.py math; recomputed in-explorer for
folders that pre-date the CSV column) displayed simply as "Hypervolume", with
an optional overlay showing the reference point — and ONE Pareto contribution,
the strict exact-match count against the universal front. The legacy per-pair
HV and TotalSolutions columns are still parsed (frozen CSV schema) but never
displayed; the near-tie seed-collaboration scoreboard is likewise kept out of
the GUI and Excel export (it remains in the Claude bundle, which documents it).

PowerCap mode (PowerCeiling campaigns, detected from the _PC<N> arm labels):

  * A "Power cap" dropdown scopes the whole GUI to one cap tier (Uncapped /
    90% / 60% / 30%, with the derived kW): the 7 BASE arms, the tier's own
    universal front (toggleable like any algorithm), and every indicator/table
    computed strictly within the tier — reference front and normalization
    bounds from that tier's arms only. The old all-arms global universal stays
    available as an off-by-default dashed overlay.
  * "Compare algorithm across caps" overlays one base arm's four tier fronts
    (traffic-light tier colors: Uncapped dark green + dashed, then light
    green / yellow / red loose -> tight); its HV bars use the run-wide frame
    so they are comparable across caps.
  * A Feasibility tab plots feasibility_summary.csv against the calibration
    targets.
  * Per-tier numbers come from the native scenario_N_*_by_cap.csv files when
    the folder has them (written by ParetoAnalyzer.analyzeScenarioByTier);
    older folders are recomputed in-explorer with the same math (the status
    bar shows which source is active).

Headless bulk export (no GUI, useful for scripting):

    python scripts/results_explorer.py <folder> --export-all <outdir>
    python scripts/results_explorer.py <folder> --export-xlsx tables.xlsx
    python scripts/results_explorer.py <folder> --claude-zip bundle.zip

Only needs pandas + numpy + matplotlib (same as the rest of the pipeline);
tkinter is imported only when the GUI actually starts.
"""

import argparse
import gzip
import json
import math
import os
import re
import sys
import tempfile
import zipfile
from datetime import datetime, timezone

import numpy as np
import pandas as pd

import matplotlib as mpl
from matplotlib.figure import Figure

mpl.rcParams.update({
    'font.size': 10,
    'axes.titlesize': 12,
    'axes.labelsize': 11,
    'axes.grid': True,
    'grid.alpha': 0.3,
    'figure.autolayout': False,
})

UNIVERSAL_KEY = 'Universal_Pareto'
UNIVERSAL_COLOR = '#000000'
UNIVERSAL_LABEL = 'Universal Pareto'

# =============================================================================
# STYLES — Okabe-Ito palette, family-stable markers (GA=diamond, SA=square,
# NSGA-II=circle, SPEA-II=triangle, AMOSA=inverted triangle). Colors are only
# DEFAULTS: the GUI lets the user recolor any algorithm on the fly; markers
# stay fixed so a family keeps its shape across recolorings.
# =============================================================================

OKABE_ITO = {
    'blue': '#0072B2', 'vermillion': '#D55E00', 'bluish_green': '#009E73',
    'reddish_purple': '#CC79A7', 'orange': '#E69F00', 'sky_blue': '#56B4E9',
    'yellow': '#F0E442', 'black': '#000000',
}

# label -> (color, marker, filled, display label). Mirrors the defaults of
# scripts/plot_scenario_pareto.py so the GUI and the publication plots agree.
ALGORITHM_STYLE = {
    'GA_Makespan':              (OKABE_ITO['blue'], 'D', False, 'GA (Makespan)'),
    'GA_WaitingTime':           (OKABE_ITO['blue'], 'D', False, 'GA (Waiting Time)'),
    'GA_Energy':                (OKABE_ITO['blue'], 'D', True, 'GA (Energy)'),
    'SA_Makespan':              (OKABE_ITO['vermillion'], 's', False, 'SA (Makespan)'),
    'SA_WaitingTime':           (OKABE_ITO['vermillion'], 's', False, 'SA (Waiting Time)'),
    'SA_Energy':                (OKABE_ITO['vermillion'], 's', True, 'SA (Energy)'),
    'GA_Makespan_Dominance':    ('#56B4E9', 'D', True, 'GA Dominance (Makespan)'),
    'GA_WaitingTime_Dominance': ('#56B4E9', 'D', True, 'GA Dominance (Waiting Time)'),
    'GA_Energy_Dominance':      ('#1B3A6B', 'D', True, 'GA Dominance (Energy)'),
    'SA_Makespan_Dominance':    ('#FF6B9D', 's', True, 'SA Dominance (Makespan)'),
    'SA_WaitingTime_Dominance': ('#FF6B9D', 's', True, 'SA Dominance (Waiting Time)'),
    'SA_Energy_Dominance':      ('#8B0000', 's', True, 'SA Dominance (Energy)'),
    'NSGA-II':                  (OKABE_ITO['bluish_green'], 'o', True, 'NSGA-II'),
    'SPEA-II':                  (OKABE_ITO['reddish_purple'], '^', True, 'SPEA-II'),
    'AMOSA':                    (OKABE_ITO['orange'], 'v', True, 'AMOSA'),
}

FALLBACK_COLORS = [
    '#0072B2', '#E69F00', '#009E73', '#CC79A7', '#56B4E9', '#D55E00',
    '#8C564B', '#7F7F7F', '#BCBD22', '#17BECF', '#9467BD', '#2CA02C',
]

# Family-stable marker inference for labels missing from ALGORITHM_STYLE.
FAMILY_MARKERS = [
    (re.compile(r'^GA[_-]', re.I), 'D'),
    (re.compile(r'^SA[_-]', re.I), 's'),
    (re.compile(r'^NSGA', re.I), 'o'),
    (re.compile(r'^SPEA', re.I), '^'),
    (re.compile(r'^AMOSA', re.I), 'v'),
    # Heuristic baselines read as one "reference" family.
    (re.compile(r'^(LPT|WorkloadAware|EnergyAware|RoundRobin|Random|MinMin|MaxMin)', re.I), 'P'),
]
FALLBACK_MARKERS = ['o', 's', 'D', '^', 'v', 'P', 'X', '*']


def family_marker(label):
    for rx, marker in FAMILY_MARKERS:
        if rx.search(label):
            return marker
    return None


def resolve_styles(algorithms):
    """Encounter-ordered algorithm labels -> {label: {color, marker, filled,
    display}}. Deterministic, so the same run always loads the same way."""
    styles = {}
    fallback_i = 0
    for algo in algorithms:
        if algo in styles or algo == UNIVERSAL_KEY:
            continue
        if algo in ALGORITHM_STYLE:
            color, marker, filled, display = ALGORITHM_STYLE[algo]
        else:
            fam = family_marker(algo)
            color = FALLBACK_COLORS[fallback_i % len(FALLBACK_COLORS)]
            marker = fam or FALLBACK_MARKERS[fallback_i % len(FALLBACK_MARKERS)]
            filled, display = True, algo
            fallback_i += 1
        styles[algo] = {'color': color, 'marker': marker,
                        'filled': filled, 'display': display}
    return styles


# =============================================================================
# DATA LAYER
# =============================================================================

def _non_dominated(points):
    """Non-dominated subset of an (n,2) array, both objectives minimized."""
    pts = np.asarray(points, dtype=float)
    keep = []
    for i, p in enumerate(pts):
        dominated = False
        for j, q in enumerate(pts):
            if i == j:
                continue
            if q[0] <= p[0] and q[1] <= p[1] and (q[0] < p[0] or q[1] < p[1]):
                dominated = True
                break
        if not dominated:
            keep.append(i)
    return pts[keep]


# =============================================================================
# POWERCAP TIERS — label parsing + indicator math ports
#
# A PowerCeiling campaign carries 4 arm sets per scenario: the 7 base arms
# uncapped plus each re-run as a constrained _PC<targetPercent> variant under
# the derived caps (CampaignRunner Phase 2). Folders written after the Java-side
# fix carry native per-tier files (scenario_N_*_by_cap.csv, computed by
# ParetoAnalyzer.analyzeScenarioByTier); for older folders the explorer
# recomputes the same numbers from the point clouds with the faithful ports
# below (ParetoAnalyzer / PerformanceMetrics / recompute_hv.py math).
# =============================================================================

TIER_RE = re.compile(r'_PC(\d+)$')
UNCAPPED_TIER = 'Uncapped'
CONTRIB_REL_EPS = 3e-3    # ParetoAnalyzer.CONTRIB_REL_EPS (near-tie credit only)
CONTRIB_ABS_FLOOR = 1e-9  # ParetoAnalyzer.CONTRIB_ABS_FLOOR
DEDUP_EPS = 1e-9          # ParetoAnalyzer.DEDUP_EPS


def split_tier(label):
    """'GA_Energy_Dominance_PC60' -> ('GA_Energy_Dominance', 'PC60');
    labels without the campaign tier suffix -> (label, 'Uncapped').
    The legacy FinalExperiment suffix form '_PC_<n>kW' is NOT a tier."""
    m = TIER_RE.search(label)
    return (label[:m.start()], 'PC' + m.group(1)) if m else (label, UNCAPPED_TIER)


def tier_sort_key(tier):
    """Uncapped first, then descending target percent (PC90, PC60, PC30)."""
    return (0, 0) if tier == UNCAPPED_TIER else (1, -int(tier[2:]))


def tier_display(tier, cap_watts=None):
    if tier == UNCAPPED_TIER:
        return 'Uncapped'
    label = f'Cap {tier[2:]}%'
    if cap_watts is not None and np.isfinite(cap_watts):
        label += f' (≈{cap_watts / 1000.0:.1f} kW)'
    return label


# ---- ParetoAnalyzer ports (raw-space union / contribution machinery) --------

def _dominates(a, b):
    """Weak Pareto dominance, both objectives minimized (ParetoAnalyzer.dominates)."""
    return a[0] <= b[0] and a[1] <= b[1] and (a[0] < b[0] or a[1] < b[1])


def _filter_nondominated(points):
    """Equivalent of ParetoAnalyzer.filterToNonDominated / nonDominatedUnion:
    non-dominated subset (weak dominance, minimization), de-duplicated at 1e-9,
    sorted by x. Implemented as an O(n log n) skyline sweep — value-identical to
    the Java O(n^2) scan (for 2-D minimization the x-sorted strictly-descending-y
    staircase IS the non-dominated set; only the representative of sub-1e-9
    duplicate clusters can differ, which CSV precision makes exactly equal
    anyway). Returns an (n, 2) float array."""
    arr = np.asarray(points, dtype=float).reshape(-1, 2)
    if not len(arr):
        return arr
    arr = np.unique(arr, axis=0)  # lexicographic sort (x, then y) + exact dedup
    keep = []
    best_y = np.inf
    for x, y in arr:
        if y < best_y:
            keep.append((x, y))
            best_y = y
    out = []
    for p in keep:  # 1e-9 near-duplicate pass (adjacent-only along the staircase)
        if out and abs(out[-1][0] - p[0]) < DEDUP_EPS and abs(out[-1][1] - p[1]) < DEDUP_EPS:
            continue
        out.append(p)
    return np.asarray(out)


def _union_bounds(points):
    """recompute_hv.py / ParetoAnalyzer.unionBounds: (idealX, idealY, nadirX, nadirY)."""
    arr = np.asarray(points, dtype=float)
    return (float(arr[:, 0].min()), float(arr[:, 1].min()),
            float(arr[:, 0].max()), float(arr[:, 1].max()))


def _normalize_union(points, bounds):
    """Union-frame min-max normalization, range floored at 1e-12."""
    ix, iy, nx, ny = bounds
    rx = max(nx - ix, 1e-12)
    ry = max(ny - iy, 1e-12)
    arr = np.asarray(points, dtype=float)
    return np.column_stack(((arr[:, 0] - ix) / rx, (arr[:, 1] - iy) / ry))


def _hv_fixed(points_norm):
    """recompute_hv.py hv_2d against (1.1, 1.1), divided by 1.21. Input is the
    union-frame-normalized point set of one run (not pre-filtered)."""
    nd = _filter_nd_sort9(points_norm)
    if not len(nd):
        return 0.0
    nd = nd[(nd[:, 0] < 1.1) & (nd[:, 1] < 1.1)]
    if not len(nd):
        return 0.0
    hv, prev_y = 0.0, 1.1
    for x, y in nd[np.argsort(nd[:, 0])]:
        hv += (1.1 - x) * (prev_y - y)
        prev_y = y
    return float(hv) / (1.1 * 1.1)


def _filter_nd_sort9(points):
    """recompute_hv.py non_dominated_sort: non-dominated subset, 9-decimal
    de-duplication, sorted by x. Returns an (n, 2) array."""
    pts = np.asarray(points, dtype=float)
    if not len(pts):
        return pts.reshape(0, 2)
    keep = [p for i, p in enumerate(pts)
            if not any(i != j and _dominates(q, p) for j, q in enumerate(pts))]
    if not keep:
        return np.empty((0, 2))
    nd = np.array(keep)
    _, idx = np.unique(np.round(nd, decimals=9), axis=0, return_index=True)
    nd = nd[np.sort(idx)]
    return nd[np.argsort(nd[:, 0])]


def _contribution_count(run_points, universal):
    """ParetoAnalyzer.paretoContributionCount: distinct universal points matched
    within 1e-9 absolute on both raw objectives (each run point credits its
    FIRST matching universal point, mirroring the Java break)."""
    run = np.asarray(run_points, dtype=float).reshape(-1, 2)
    uni = np.asarray(universal, dtype=float).reshape(-1, 2)
    if not len(run) or not len(uni):
        return 0
    hit = ((np.abs(run[:, None, 0] - uni[None, :, 0]) < DEDUP_EPS)
           & (np.abs(run[:, None, 1] - uni[None, :, 1]) < DEDUP_EPS))
    any_hit = hit.any(axis=1)
    first = hit.argmax(axis=1)  # index of the first True per row
    return len(set(first[any_hit].tolist()))


def _union_contribution_count(runs_points, universal):
    """ParetoAnalyzer.unionContributionCount: distinct universal points matched
    by ANY of the label's runs (the MEAN-row integer)."""
    stacked = [np.asarray(p, dtype=float).reshape(-1, 2)
               for p in runs_points if len(p)]
    if not stacked:
        return 0
    return _contribution_count(np.vstack(stacked), universal)


def _near_tie_credit(run_points, universal):
    """ParetoAnalyzer.paretoContributionCountEps: distinct universal points some
    run point near-ties (CONTRIB_REL_EPS relative on BOTH objectives, tolerance
    scaled to the universal point's own magnitude, nearest-match)."""
    run = np.asarray(run_points, dtype=float).reshape(-1, 2)
    uni = np.asarray(universal, dtype=float).reshape(-1, 2)
    if not len(run) or not len(uni):
        return 0
    tol = np.maximum(np.abs(uni) * CONTRIB_REL_EPS, CONTRIB_ABS_FLOOR)  # (m, 2)
    d = np.abs(run[:, None, :] - uni[None, :, :])                       # (n, m, 2)
    within = (d <= tol[None, :, :]).all(axis=2)                         # (n, m)
    score = ((d / tol[None, :, :]) ** 2).sum(axis=2)                    # (n, m)
    score[~within] = np.inf
    best = score.argmin(axis=1)
    ok = within.any(axis=1)
    return len(set(best[ok].tolist()))


# ---- PerformanceMetrics ports (legacy per-pair-frame HV/GD/IGD/Spacing) -----

def _legacy_pair_metrics(seed_front, universal):
    """Port of ParetoAnalyzer.computeLegacyIndicators: PerformanceMetrics on the
    pair {seed non-dominated front, universal front} (reference index 1) in the
    PAIR's own min-max frame. Returns (hv, gd, igd, spacing) with the runner
    fallbacks (hv=0, gd=igd=MAX, spacing=0) on degenerate input."""
    if universal is None or not len(universal):
        return 0.0, sys.float_info.max, sys.float_info.max, 0.0
    seed = [list(map(float, p)) for p in seed_front]
    n_seed = len(seed)
    if not seed:
        seed = [[sys.float_info.max, sys.float_info.max]]  # runner dummy point
    univ = [list(map(float, p)) for p in universal]
    try:
        fronts = [sorted(seed, key=lambda p: p[0]), sorted(univ, key=lambda p: p[0])]
        # findMinMax: first/last elements only (fronts are x-sorted staircases).
        f1min = f1max = fronts[0][0][0]
        f2min = f2max = fronts[0][0][1]
        for fr in fronts:
            first, last = fr[0], fr[-1]
            f1min = min(f1min, first[0])
            f1max = max(f1max, last[0])
            f2min = min(f2min, last[1])
            f2max = max(f2max, first[1])
        r1, r2 = f1max - f1min, f2max - f2min
        norm = [[[((p[0] - f1min) / r1) if r1 > 0 else 0.0,
                  ((p[1] - f2min) / r2) if r2 > 0 else 0.0] for p in fr]
                for fr in fronts]
        ns, nu = norm[0], norm[1]

        hv = abs(1.0 - ns[0][0]) * abs(1.0 - ns[0][1])
        for i in range(1, len(ns)):
            hv += abs(1.0 - ns[i][0]) * abs(ns[i - 1][1] - ns[i][1])

        a, u = np.asarray(ns, dtype=float), np.asarray(nu, dtype=float)
        d = np.sqrt(((a[:, None, :] - u[None, :, :]) ** 2).sum(axis=2))
        gd = float(d.min(axis=1).mean())
        igd = float(d.min(axis=0).mean())

        spacing = 0.0
        if n_seed > 1:
            p = ns
            dist = [abs(p[1][0] - p[0][0]) + abs(p[0][1] - p[1][1])]
            prev = dist[0]
            for i in range(1, len(p) - 1):
                nxt = abs(p[i + 1][0] - p[i][0]) + abs(p[i][1] - p[i + 1][1])
                dist.append(min(prev, nxt))
                prev = nxt
            last = len(p) - 1
            dist.append(p[last][0] - p[last - 1][0] + p[last - 1][1] - p[last][1])
            d_avg = sum(dist) / len(dist)
            spacing = math.sqrt(sum((d_avg - d) ** 2 for d in dist) / (len(p) - 1))
        return float(hv), float(gd), float(igd), float(spacing)
    except Exception:
        return 0.0, sys.float_info.max, sys.float_info.max, 0.0


def _legacy_own_frame_hv(universal):
    """ParetoAnalyzer.universalHV: the universal front's legacy HV in its OWN
    min-max frame (comparable to nothing; kept for the trailer row)."""
    if universal is None or not len(universal):
        return 0.0
    try:
        fr = sorted([list(map(float, p)) for p in universal], key=lambda p: p[0])
        f1min, f1max = fr[0][0], fr[-1][0]
        f2min, f2max = fr[-1][1], fr[0][1]
        r1, r2 = f1max - f1min, f2max - f2min
        ns = [[((p[0] - f1min) / r1) if r1 > 0 else 0.0,
               ((p[1] - f2min) / r2) if r2 > 0 else 0.0] for p in fr]
        hv = abs(1.0 - ns[0][0]) * abs(1.0 - ns[0][1])
        for i in range(1, len(ns)):
            hv += abs(1.0 - ns[i][0]) * abs(ns[i - 1][1] - ns[i][1])
        return float(hv)
    except Exception:
        return 0.0


class TierData:
    """One cap tier's slice of a scenario: fronts, universal front and per-tier
    indicator tables, keyed by BASE algorithm labels."""

    def __init__(self, tier):
        self.tier = tier
        self.cap_watts = None       # derived cap (W); None for Uncapped
        self.points = None          # tier's cloud rows (Algorithm = base label)
        self.universal = None       # tier universal front DataFrame
        self.fronts = {}            # base label -> pooled front DataFrame
        self.metrics_seed = None
        self.metrics_mean = None
        self.metrics_std = None
        self.universal_metrics = {}
        self.collab_seed = None
        self.collab_mean = None
        self.collab_std = None
        self.collab_universal = {}  # {'mean': {...}, 'std': {...}}


class ScenarioData:
    """All artifacts of one scenario_N_* set, parsed and split."""

    def __init__(self, number, name):
        self.number = number
        self.name = name
        self.obj_names = []          # [obj_x, obj_y] as named in the CSV header
        self.points = None           # raw per-seed cloud rows (no universal rows)
        self.universal = None        # pooled universal front DataFrame
        self.fronts = {}             # label -> per-algorithm pooled front DataFrame
        self.metrics_seed = None     # per-seed metric rows (numeric Seed)
        self.metrics_mean = None     # MEAN rows, indexed by Algorithm
        self.metrics_std = None      # STDDEV rows, indexed by Algorithm
        self.universal_metrics = {}  # trailer row of performance_metrics.csv
        self.has_hv_fixed = False
        self.collab_seed = None      # per-seed collaboration rows (or None)
        self.collab_mean = None
        self.collab_std = None
        self.collab_universal = {}   # {'mean': {...}, 'std': {...}}
        self.details = None          # parsed scenario_N_solution_details.json
        self.tiers = {}              # tier name -> TierData (powercap folders only)


class ExperimentData:
    """A loaded campaign result folder."""

    def __init__(self, folder):
        self.folder = os.path.abspath(folder)
        self.name = os.path.basename(self.folder.rstrip('/'))
        self.scenarios = {}      # number -> ScenarioData
        self.algorithms = []     # encounter-ordered labels (no Universal_Pareto)
        self.plot_options = {}
        self.hv_fixed_recomputed = False  # True when any scenario's HV_fixed
                                          # was computed in-explorer (folder
                                          # pre-dates the CSV column)
        # PowerCap mode (set by _load_powercap when _PC<N> arms are present):
        self.is_powercap = False
        self.base_algorithms = []  # encounter-ordered base labels (7 arms)
        self.tier_names = []       # ordered: Uncapped, PC90, PC60, PC30
        self.cap_watts = {}        # tier -> derived cap (W), when known
        self.feasibility = None    # feasibility_summary.csv DataFrame (or None)
        self.by_cap_source = None  # 'native' | 'recomputed' | None

    # ---- loading --------------------------------------------------------

    @classmethod
    def load(cls, folder):
        exp = cls(folder)
        nums = sorted(
            int(m.group(1))
            for f in os.listdir(exp.folder)
            if (m := re.match(r'scenario_(\d+)_pareto_graph_data\.csv$', f)))
        if not nums:
            raise ValueError(
                f'No scenario_N_pareto_graph_data.csv found in {exp.folder} — '
                'not a campaign result folder.')

        scen_names = exp._load_scenario_names()
        po = os.path.join(exp.folder, 'plot_options.json')
        if os.path.isfile(po):
            try:
                with open(po) as fh:
                    exp.plot_options = json.load(fh)
            except (OSError, json.JSONDecodeError):
                pass

        for n in nums:
            scn = ScenarioData(n, scen_names.get(n, f'Scenario {n}'))
            exp._load_points(scn)
            exp._load_fronts(scn)
            exp._load_metrics(scn)
            exp._ensure_hv_fixed(scn)
            exp._load_collaboration(scn)
            exp._load_details(scn)
            exp.scenarios[n] = scn

        # Encounter-ordered algorithm list from metrics (falls back to points).
        seen = {}
        for scn in exp.scenarios.values():
            src = scn.metrics_seed if scn.metrics_seed is not None else scn.points
            if src is None:
                continue
            for a in src['Algorithm']:
                if a != UNIVERSAL_KEY:
                    seen.setdefault(a, None)
        exp.algorithms = list(seen)
        exp._load_powercap()
        return exp

    def _path(self, fname):
        return os.path.join(self.folder, fname)

    def _load_scenario_names(self):
        names = {}
        path = self._path('experiment_summary.csv')
        if os.path.isfile(path):
            try:
                df = pd.read_csv(path, usecols=['Scenario', 'ScenarioName'])
                for _, row in df.drop_duplicates('Scenario').iterrows():
                    names[int(row['Scenario'])] = str(row['ScenarioName']).replace('_', ' ')
            except (ValueError, KeyError, OSError):
                pass
        return names

    def _load_points(self, scn):
        df = pd.read_csv(self._path(f'scenario_{scn.number}_pareto_graph_data.csv'),
                         dtype={'Seed': str})
        aux = {'Algorithm', 'Seed', 'IsUniversalPareto', 'PeakPowerWatts'}
        scn.obj_names = [c for c in df.columns if c not in aux][:2]
        for c in scn.obj_names:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        is_univ_row = df['Algorithm'] == UNIVERSAL_KEY
        scn.universal = (df[is_univ_row]
                         .sort_values(scn.obj_names[0])
                         .reset_index(drop=True))
        scn.points = df[~is_univ_row].reset_index(drop=True)

    def _load_fronts(self, scn):
        path = self._path(f'scenario_{scn.number}_algorithm_pareto_fronts.csv')
        if os.path.isfile(path):
            df = pd.read_csv(path)
            obj = [c for c in df.columns if c != 'Algorithm'][:2]
            for label, grp in df.groupby('Algorithm', sort=False):
                scn.fronts[label] = (grp.rename(columns=dict(zip(obj, scn.obj_names)))
                                     .sort_values(scn.obj_names[0])
                                     .reset_index(drop=True))
        else:
            # Older folders: derive each algorithm's pooled front from the clouds.
            for label, grp in scn.points.groupby('Algorithm', sort=False):
                nd = _non_dominated(grp[scn.obj_names].to_numpy())
                nd = nd[np.argsort(nd[:, 0])]
                scn.fronts[label] = pd.DataFrame(nd, columns=scn.obj_names)

    def _load_metrics(self, scn):
        path = self._path(f'scenario_{scn.number}_performance_metrics.csv')
        if not os.path.isfile(path):
            return
        df = pd.read_csv(path, dtype={'Seed': str})
        scn.has_hv_fixed = 'HV_fixed' in df.columns
        num_cols = [c for c in df.columns if c not in ('Algorithm', 'Seed')]
        for c in num_cols:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        is_univ = df['Algorithm'] == UNIVERSAL_KEY
        seed_mask = df['Seed'].str.fullmatch(r'\d+') & ~is_univ
        scn.metrics_seed = df[seed_mask].copy()
        scn.metrics_seed['Seed'] = scn.metrics_seed['Seed'].astype(int)
        scn.metrics_mean = df[(df['Seed'] == 'MEAN') & ~is_univ].set_index('Algorithm')
        scn.metrics_std = df[(df['Seed'] == 'STDDEV') & ~is_univ].set_index('Algorithm')
        univ = df[is_univ]
        if len(univ):
            scn.universal_metrics = univ.iloc[0].to_dict()

    def _ensure_hv_fixed(self, scn):
        """Folders written before PR #220 have no HV_fixed column. Since the
        GUI shows HV_fixed as its ONE hypervolume, compute it here with the
        exact recompute_hv.py math (fixed reference 1.1/1.1 in the scenario's
        pooled ideal/nadir frame — the same port the powercap recompute uses),
        so every folder gets the same single, comparable Hypervolume."""
        if scn.metrics_seed is None or scn.has_hv_fixed or not len(scn.metrics_seed):
            return
        objs = scn.obj_names
        all_pts = [scn.points[objs].to_numpy(dtype=float)]
        if scn.universal is not None and len(scn.universal):
            all_pts.append(scn.universal[objs].to_numpy(dtype=float))
        bounds = _union_bounds(np.vstack(all_pts))

        hv_by_run = {}
        for (algo, seed), grp in scn.points.groupby(['Algorithm', 'Seed'], sort=False):
            pts = grp[objs].to_numpy(dtype=float)
            hv_by_run[(algo, int(seed))] = _hv_fixed(_normalize_union(pts, bounds))
        scn.metrics_seed['HV_fixed'] = [
            hv_by_run.get((r['Algorithm'], int(r['Seed'])), np.nan)
            for _, r in scn.metrics_seed.iterrows()]

        # MEAN/STDDEV per algorithm (population stddev, matching the Java rows).
        grouped = scn.metrics_seed.groupby('Algorithm', sort=False)['HV_fixed']
        if scn.metrics_mean is not None:
            scn.metrics_mean['HV_fixed'] = grouped.mean()
        if scn.metrics_std is not None:
            scn.metrics_std['HV_fixed'] = grouped.std(ddof=0)
        if scn.universal_metrics and scn.universal is not None and len(scn.universal):
            scn.universal_metrics['HV_fixed'] = _hv_fixed(
                _normalize_union(scn.universal[objs].to_numpy(dtype=float), bounds))
        scn.has_hv_fixed = True
        self.hv_fixed_recomputed = True

    def _load_collaboration(self, scn):
        path = self._path(f'scenario_{scn.number}_seed_collaboration.csv')
        if not os.path.isfile(path):
            return
        df = pd.read_csv(path, dtype={'Seed': str})
        num_cols = [c for c in df.columns if c not in ('Algorithm', 'Seed')]
        for c in num_cols:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        is_univ = df['Algorithm'] == UNIVERSAL_KEY
        seed_mask = df['Seed'].str.fullmatch(r'\d+') & ~is_univ
        scn.collab_seed = df[seed_mask].copy()
        scn.collab_seed['Seed'] = scn.collab_seed['Seed'].astype(int)
        scn.collab_mean = df[(df['Seed'] == 'MEAN') & ~is_univ].set_index('Algorithm')
        scn.collab_std = df[(df['Seed'] == 'STDDEV') & ~is_univ].set_index('Algorithm')
        for kind in ('MEAN', 'STDDEV'):
            row = df[is_univ & (df['Seed'] == kind)]
            if len(row):
                scn.collab_universal[kind.lower()] = row.iloc[0].to_dict()

    def _load_details(self, scn):
        for suffix, opener in ((".json", open), (".json.gz", gzip.open)):
            path = self._path(f'scenario_{scn.number}_solution_details{suffix}')
            if os.path.isfile(path):
                try:
                    with opener(path, 'rt') as fh:
                        scn.details = json.load(fh)
                except (OSError, json.JSONDecodeError) as e:
                    print(f'warning: could not parse {path}: {e}', file=sys.stderr)
                return

    # ---- powercap (cap-tier) loading -------------------------------------

    def _load_powercap(self):
        """Detects a PowerCeiling folder (arm labels with _PC<N> suffixes) and
        builds per-tier data for every scenario: native *_by_cap.csv files when
        the run was produced by a build that writes them, else a faithful
        in-explorer recompute from the point clouds. No-op for other studies."""
        bases, tiers = {}, {}
        for a in self.algorithms:
            b, t = split_tier(a)
            bases.setdefault(b, None)
            tiers.setdefault(t, None)
        self.is_powercap = any(t != UNCAPPED_TIER for t in tiers)
        if not self.is_powercap:
            return
        self.base_algorithms = list(bases)
        self.tier_names = sorted(tiers, key=tier_sort_key)
        self._load_feasibility()

        native = all(
            os.path.isfile(self._path(f'scenario_{n}_performance_metrics_by_cap.csv'))
            for n in self.scenarios)
        self.by_cap_source = 'native' if native else 'recomputed'
        for scn in self.scenarios.values():
            self._slice_tier_points(scn)
            if native:
                self._load_tiers_native(scn)
            else:
                self._recompute_tiers(scn)
        self._resolve_cap_watts()

    def _full_label(self, base, tier):
        return base if tier == UNCAPPED_TIER else f'{base}_{tier}'

    def _slice_tier_points(self, scn):
        """Splits the clouds and pooled fronts by tier, relabelling arms to
        their base names (tier identity lives in the TierData)."""
        tier_of = {a: split_tier(a)[1] for a in self.algorithms}
        base_of = {a: split_tier(a)[0] for a in self.algorithms}
        for tier in self.tier_names:
            td = TierData(tier)
            mask = scn.points['Algorithm'].map(lambda a: tier_of.get(a) == tier)
            pts = scn.points[mask].copy()
            pts['Algorithm'] = pts['Algorithm'].map(base_of)
            td.points = pts.reset_index(drop=True)
            for label, front in scn.fronts.items():
                if tier_of.get(label) == tier:
                    td.fronts[base_of[label]] = front
            scn.tiers[tier] = td

    def _load_feasibility(self):
        path = self._path('feasibility_summary.csv')
        if os.path.isfile(path):
            try:
                self.feasibility = pd.read_csv(path)
            except (OSError, ValueError):
                self.feasibility = None

    def _resolve_cap_watts(self):
        """Tier -> derived cap Watts: native CapWatts columns first, else the
        distinct CapWatts of feasibility_summary.csv matched to the PC tiers in
        descending order (looser target = higher cap by construction)."""
        watts = {}
        for scn in self.scenarios.values():
            for tier, td in scn.tiers.items():
                if td.cap_watts is not None and np.isfinite(td.cap_watts):
                    watts.setdefault(tier, float(td.cap_watts))
        pc_tiers = [t for t in self.tier_names if t != UNCAPPED_TIER]
        if not watts and self.feasibility is not None and 'CapWatts' in self.feasibility:
            caps = sorted(pd.unique(self.feasibility['CapWatts'].dropna()), reverse=True)
            watts = {t: float(w) for t, w in zip(pc_tiers, caps)}
        self.cap_watts = watts
        for scn in self.scenarios.values():
            for tier, td in scn.tiers.items():
                td.cap_watts = watts.get(tier)

    # ---- native *_by_cap.csv loaders --------------------------------------

    def _load_tiers_native(self, scn):
        n = scn.number
        uni_path = self._path(f'scenario_{n}_universal_fronts_by_cap.csv')
        if os.path.isfile(uni_path):
            df = pd.read_csv(uni_path)
            obj = [c for c in df.columns if c not in ('CapTier', 'CapWatts')][:2]
            for tier, grp in df.groupby('CapTier', sort=False):
                td = scn.tiers.get(tier)
                if td is None:
                    continue
                td.universal = (grp[obj]
                                .rename(columns=dict(zip(obj, scn.obj_names)))
                                .sort_values(scn.obj_names[0])
                                .reset_index(drop=True))
                if 'CapWatts' in grp and grp['CapWatts'].notna().any():
                    td.cap_watts = float(grp['CapWatts'].dropna().iloc[0])

        df = pd.read_csv(self._path(f'scenario_{n}_performance_metrics_by_cap.csv'),
                         dtype={'Seed': str})
        num_cols = [c for c in df.columns
                    if c not in ('CapTier', 'Algorithm', 'Seed')]
        for c in num_cols:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        for tier, grp in df.groupby('CapTier', sort=False):
            td = scn.tiers.get(tier)
            if td is None:
                continue
            if 'CapWatts' in grp and grp['CapWatts'].notna().any():
                td.cap_watts = float(grp['CapWatts'].dropna().iloc[0])
            grp = grp.drop(columns=[c for c in ('CapTier', 'CapWatts') if c in grp])
            is_univ = grp['Algorithm'] == UNIVERSAL_KEY
            seed_mask = grp['Seed'].str.fullmatch(r'\d+') & ~is_univ
            td.metrics_seed = grp[seed_mask].copy()
            td.metrics_seed['Seed'] = td.metrics_seed['Seed'].astype(int)
            td.metrics_mean = grp[(grp['Seed'] == 'MEAN') & ~is_univ].set_index('Algorithm')
            td.metrics_std = grp[(grp['Seed'] == 'STDDEV') & ~is_univ].set_index('Algorithm')
            univ = grp[is_univ]
            if len(univ):
                td.universal_metrics = univ.iloc[0].to_dict()

        collab_path = self._path(f'scenario_{n}_seed_collaboration_by_cap.csv')
        if os.path.isfile(collab_path):
            df = pd.read_csv(collab_path, dtype={'Seed': str})
            num_cols = [c for c in df.columns if c not in ('CapTier', 'Algorithm', 'Seed')]
            for c in num_cols:
                df[c] = pd.to_numeric(df[c], errors='coerce')
            for tier, grp in df.groupby('CapTier', sort=False):
                td = scn.tiers.get(tier)
                if td is None:
                    continue
                grp = grp.drop(columns=['CapTier'])
                is_univ = grp['Algorithm'] == UNIVERSAL_KEY
                seed_mask = grp['Seed'].str.fullmatch(r'\d+') & ~is_univ
                td.collab_seed = grp[seed_mask].copy()
                td.collab_seed['Seed'] = td.collab_seed['Seed'].astype(int)
                td.collab_mean = grp[(grp['Seed'] == 'MEAN') & ~is_univ].set_index('Algorithm')
                td.collab_std = grp[(grp['Seed'] == 'STDDEV') & ~is_univ].set_index('Algorithm')
                for kind in ('MEAN', 'STDDEV'):
                    row = grp[is_univ & (grp['Seed'] == kind)]
                    if len(row):
                        td.collab_universal[kind.lower()] = row.iloc[0].to_dict()

    # ---- in-explorer recompute (pre-fix folders) ---------------------------

    def _recompute_tiers(self, scn):
        """Recomputes every per-tier table exactly the way the Java side does
        (ParetoAnalyzer.analyzeScenarioByTier + ExperimentReporter *_by_cap
        writers): tier universal front, tier union bounds, per-seed legacy
        HV/GD/IGD/Spacing (per-pair frame vs the TIER universal), HV_fixed
        (tier frame), strict contribution counts and the near-tie
        seed-collaboration scoreboard. TimeMs comes from the run's own row in
        the global performance_metrics.csv (runtime is per run, not re-derived)."""
        metric_cols = ['HV', 'GD', 'IGD', 'Spacing', 'NonDomSolutions',
                       'TotalSolutions', 'ParetoContribution', 'TimeMs', 'HV_fixed']
        time_lookup = {}
        if scn.metrics_seed is not None and 'TimeMs' in scn.metrics_seed.columns:
            for _, r in scn.metrics_seed.iterrows():
                time_lookup[(r['Algorithm'], int(r['Seed']))] = r['TimeMs']

        for tier, td in scn.tiers.items():
            if td.points is None or not len(td.points):
                continue
            objs = scn.obj_names
            all_pts = td.points[objs].to_numpy(dtype=float)
            universal = _filter_nondominated(all_pts)
            td.universal = pd.DataFrame(universal, columns=objs)
            bounds = _union_bounds(np.vstack([all_pts, np.asarray(universal)]))

            runs = []  # (base, seed:int, run_pts) in cloud row order
            for (base, seed), grp in td.points.groupby(['Algorithm', 'Seed'], sort=False):
                runs.append((base, int(seed), grp[objs].to_numpy(dtype=float)))

            rows, mean_rows, std_rows = [], [], []
            by_label = {}
            for base, seed, pts in runs:
                by_label.setdefault(base, []).append((seed, pts))
            for base, seed_runs in by_label.items():
                label_rows = []
                for seed, pts in seed_runs:
                    nd = _filter_nondominated(pts)
                    hv, gd, igd, spacing = _legacy_pair_metrics(nd, universal)
                    row = {
                        'Algorithm': base, 'Seed': seed,
                        'HV': hv, 'GD': gd, 'IGD': igd, 'Spacing': spacing,
                        'NonDomSolutions': len(nd), 'TotalSolutions': len(pts),
                        'ParetoContribution': _contribution_count(pts, universal),
                        'TimeMs': time_lookup.get(
                            (self._full_label(base, tier), seed), np.nan),
                        'HV_fixed': _hv_fixed(_normalize_union(pts, bounds)),
                    }
                    label_rows.append(row)
                    rows.append(row)
                lr = pd.DataFrame(label_rows)
                times = lr['TimeMs'].dropna()
                times = times[times > 0]
                mean_rows.append({
                    'Algorithm': base,
                    'HV': lr['HV'].mean(), 'GD': lr['GD'].mean(),
                    'IGD': lr['IGD'].mean(), 'Spacing': lr['Spacing'].mean(),
                    'NonDomSolutions': lr['NonDomSolutions'].mean(),
                    'TotalSolutions': lr['TotalSolutions'].mean(),
                    'ParetoContribution': _union_contribution_count(
                        [pts for _, pts in seed_runs], universal),
                    'TimeMs': int(times.sum() // len(times)) if len(times) else 0,
                    'HV_fixed': lr['HV_fixed'].mean(),
                })
                std_rows.append({
                    'Algorithm': base,
                    'HV': lr['HV'].std(ddof=0), 'GD': lr['GD'].std(ddof=0),
                    'IGD': lr['IGD'].std(ddof=0), 'Spacing': lr['Spacing'].std(ddof=0),
                    'HV_fixed': lr['HV_fixed'].std(ddof=0),
                })
            td.metrics_seed = pd.DataFrame(rows, columns=['Algorithm', 'Seed'] + metric_cols)
            td.metrics_mean = pd.DataFrame(mean_rows).set_index('Algorithm')
            td.metrics_std = pd.DataFrame(std_rows).set_index('Algorithm')
            td.universal_metrics = {
                'Algorithm': UNIVERSAL_KEY, 'Seed': 'ALL',
                'HV': _legacy_own_frame_hv(universal),
                'GD': 0.0, 'IGD': 0.0, 'Spacing': 0.0,
                'NonDomSolutions': len(universal), 'TotalSolutions': len(universal),
                'ParetoContribution': len(universal), 'TimeMs': 0,
                'HV_fixed': _hv_fixed(_normalize_union(universal, bounds)),
            }
            self._recompute_tier_collab(td, objs, bounds, by_label)

    def _recompute_tier_collab(self, td, objs, bounds, by_label):
        """Per-seed all-arms-within-tier universal fronts + near-tie credit
        (ParetoAnalyzer.analyzeSeedCollaboration semantics)."""
        seed_pool = {}  # seed -> stacked tier points in cloud row order
        for _, grp in td.points.groupby('Seed', sort=False):
            seed_pool[int(grp['Seed'].iloc[0])] = grp[objs].to_numpy(dtype=float)
        seed_univ = {s: _filter_nondominated(p) for s, p in seed_pool.items()}
        seed_hvf = {s: _hv_fixed(_normalize_union(u, bounds)) if len(u) else np.nan
                    for s, u in seed_univ.items()}

        rows, mean_rows, std_rows = [], [], []
        for base, seed_runs in by_label.items():
            counts, pcts = [], []
            for seed, pts in seed_runs:
                univ = seed_univ.get(seed, [])
                count = _near_tie_credit(pts, univ)
                pct = 100.0 * count / len(univ) if len(univ) else np.nan
                rows.append({'Algorithm': base, 'Seed': seed,
                             'SeedUniversalFrontSize': len(univ),
                             'SeedUniversalHV_fixed': seed_hvf.get(seed, np.nan),
                             'ContributionCount': count, 'ContributionPct': pct})
                counts.append(count)
                pcts.append(pct)
            counts, pcts = np.asarray(counts, float), np.asarray(pcts, float)
            mean_rows.append({'Algorithm': base, 'ContributionCount': counts.mean(),
                              'ContributionPct': np.nanmean(pcts) if len(pcts) else np.nan})
            std_rows.append({'Algorithm': base, 'ContributionCount': counts.std(ddof=0),
                             'ContributionPct': np.nanstd(pcts) if len(pcts) else np.nan})
        td.collab_seed = pd.DataFrame(rows)
        td.collab_mean = pd.DataFrame(mean_rows).set_index('Algorithm')
        td.collab_std = pd.DataFrame(std_rows).set_index('Algorithm')
        sizes = np.asarray([len(u) for u in seed_univ.values()], float)
        hvfs = np.asarray(list(seed_hvf.values()), float)
        if len(sizes):
            td.collab_universal = {
                'mean': {'SeedUniversalFrontSize': sizes.mean(),
                         'SeedUniversalHV_fixed': np.nanmean(hvfs)},
                'std': {'SeedUniversalFrontSize': sizes.std(ddof=0),
                        'SeedUniversalHV_fixed': np.nanstd(hvfs)},
            }

    # ---- convenience ----------------------------------------------------

    @property
    def objective_names(self):
        first = next(iter(self.scenarios.values()))
        return first.obj_names

    def study_label(self):
        objs = self.objective_names
        return f'{objs[0]} vs {objs[1]}' if len(objs) == 2 else self.name

# =============================================================================
# FIGURE BUILDERS — pure functions (Figure in, no pyplot, no GUI state) so the
# same code drives the embedded canvases, the Save buttons, and --export-all.
# =============================================================================

# One hypervolume (HV_fixed, shown as just "Hypervolume"), one Pareto
# contribution (the strict count). The legacy HV / TotalSolutions columns and
# the near-tie collaboration shares are deliberately absent from every list.
METRIC_CHOICES = [
    # (key, display, ylabel)
    ('HV_fixed', 'Hypervolume', 'Hypervolume'),
    ('GD', 'Generational Distance', 'GD'),
    ('IGD', 'Inverted Generational Distance', 'IGD'),
    ('Spacing', 'Spacing', 'Spacing'),
    ('ParetoContribution', 'Pareto Contribution',
     'universal-front points contributed'),
]


def available_metrics(exp):
    """METRIC_CHOICES when the folder has metric tables at all."""
    if not any(s.metrics_seed is not None for s in exp.scenarios.values()):
        return []
    return list(METRIC_CHOICES)


# PowerCap per-cap view: same indicators, computed strictly within the selected
# tier (reference front + normalization bounds from that tier's arms only).
POWERCAP_METRIC_CHOICES = [
    ('HV_fixed', 'Hypervolume (tier frame)', 'Hypervolume (tier frame)'),
    ('GD', 'Generational Distance (tier reference)', 'GD'),
    ('IGD', 'Inverted Generational Distance (tier reference)', 'IGD'),
    ('Spacing', 'Spacing', 'Spacing'),
    ('ParetoContribution', 'Pareto Contribution (tier universal)',
     'universal-front points contributed'),
]

# PowerCap compare mode: indicators that make sense ACROSS tiers for one arm.
# Hypervolume uses the run-wide frame (the stored global HV_fixed — pooled
# ideal/nadir over all tiers), so bars are directly comparable across caps;
# GD/IGD are referenced to each tier's own universal front ("competitiveness
# at that cap"). These carry a 4th element: the compare-mode value source.
COMPARE_METRIC_CHOICES = [
    ('HV_fixed_run', 'Hypervolume (run-wide frame, comparable across caps)',
     'Hypervolume (run-wide frame)', 'global-seed'),
    ('GD', 'Generational Distance (each tier vs its own universal)', 'GD', 'tier-seed'),
    ('IGD', 'Inverted Generational Distance (each tier vs its own universal)',
     'IGD', 'tier-seed'),
    ('ParetoContribution', 'Pareto Contribution (tier universal)',
     'universal-front points contributed', 'tier-union'),
]


def powercap_metrics(exp):
    return list(POWERCAP_METRIC_CHOICES)


def _tier_metric_values(td, algo, key, source):
    """(mean, std) of one per-tier metric for one base algorithm (mirrors
    _metric_values, on a TierData). source 'union' reads the MEAN row's
    union-over-seeds contribution count (compare mode only)."""
    if td.metrics_seed is None:
        return (np.nan, np.nan)
    if source == 'union':
        if td.metrics_mean is not None and algo in td.metrics_mean.index:
            return (td.metrics_mean.loc[algo, 'ParetoContribution'], np.nan)
        return (np.nan, np.nan)
    rows = td.metrics_seed[td.metrics_seed['Algorithm'] == algo]
    if key not in rows.columns or not len(rows):
        return (np.nan, np.nan)
    vals = rows[key].dropna()
    if not len(vals):
        return (np.nan, np.nan)
    return (vals.mean(), vals.std(ddof=1) if len(vals) > 1 else 0.0)


def _compare_metric_values(exp, scn, base, tier, key, source):
    """(mean, std) of one compare-mode metric for one (base arm, tier)."""
    td = scn.tiers.get(tier)
    if source == 'global-seed':
        # Stored global HV_fixed rows for the tier-suffixed arm: the run-wide
        # frame is exactly the pooled scenario frame, comparable across tiers.
        if scn.metrics_seed is None or 'HV_fixed' not in scn.metrics_seed.columns:
            return (np.nan, np.nan)
        full = base if tier == UNCAPPED_TIER else f'{base}_{tier}'
        vals = scn.metrics_seed.loc[
            scn.metrics_seed['Algorithm'] == full, 'HV_fixed'].dropna()
        if not len(vals):
            return (np.nan, np.nan)
        return (vals.mean(), vals.std(ddof=1) if len(vals) > 1 else 0.0)
    if td is None:
        return (np.nan, np.nan)
    if source == 'tier-union':
        return _tier_metric_values(td, base, 'ParetoContribution', 'union')
    return _tier_metric_values(td, base, key, 'seed')


def _visible(exp, opts):
    hidden = opts.get('hidden', set())
    return [a for a in exp.algorithms if a not in hidden]


def _apply_legend(ax_or_fig, opts, handles=None, labels=None):
    if not opts.get('show_legend', True):
        return
    if handles is not None:
        ax_or_fig.legend(handles, labels, fontsize=9, framealpha=0.85)
    else:
        ax_or_fig.legend(fontsize=9, framealpha=0.85)


def _bounds_from(points_df, universal_df, obj_names):
    frames = [points_df[obj_names]]
    if universal_df is not None and len(universal_df):
        frames.append(universal_df[obj_names])
    allpts = pd.concat(frames, ignore_index=True)
    out = {}
    for c in obj_names:
        ideal = float(allpts[c].min())
        span = max(float(allpts[c].max()) - ideal, 1e-12)
        out[c] = (ideal, span)
    return out


def scenario_bounds(scn):
    """Pooled per-scenario ideal/nadir over ALL algorithms' published points
    plus the universal front — the same frame HV_fixed normalizes by. Stable
    under visibility toggles so the picture doesn't rescale when algorithms
    are hidden. Returns {obj_name: (ideal, span)}."""
    return _bounds_from(scn.points, scn.universal, scn.obj_names)


def tier_bounds(td, obj_names):
    """Per-tier ideal/nadir (the frame the tier's HV_fixed uses)."""
    return _bounds_from(td.points, td.universal, obj_names)


# Tier identity colors — a traffic-light severity scale (owner-chosen):
# Uncapped dark green, then loose -> tight caps as light green / yellow / red.
# Validated for CVD separation; Uncapped additionally draws dashed and grouped
# bars carry tier order by position, so identity never rides on color alone.
# Campaigns with extra cap tiers fall back to the trailing spare hues.
TIER_COLOR_UNCAPPED = '#226B27'
TIER_COLORS_CAPPED = ['#8BC34A', '#D9A400', '#C62828', '#7B1FA2', '#00838F', '#5D4037']
GLOBAL_UNIVERSAL_COLOR = '#9AA0A6'


def tier_color(exp, tier):
    if tier == UNCAPPED_TIER:
        return TIER_COLOR_UNCAPPED
    capped = [t for t in exp.tier_names if t != UNCAPPED_TIER]
    try:
        return TIER_COLORS_CAPPED[capped.index(tier) % len(TIER_COLORS_CAPPED)]
    except ValueError:
        return TIER_COLORS_CAPPED[-1]


def _draw_ref_point(ax, bounds, ox, oy, normalized, ms):
    """Marks the Hypervolume reference point (nadir + 10% of range on both
    objectives = (1.1, 1.1) in the normalized frame) with dotted guide lines.
    `bounds` is the {obj: (ideal, span)} dict the figure normalizes by."""
    if normalized:
        rx, ry = 1.1, 1.1
    else:
        rx = bounds[ox][0] + 1.1 * bounds[ox][1]
        ry = bounds[oy][0] + 1.1 * bounds[oy][1]
    ax.axvline(rx, color='#999999', linestyle=':', linewidth=0.9, zorder=1)
    ax.axhline(ry, color='#999999', linestyle=':', linewidth=0.9, zorder=1)
    ax.plot([rx], [ry], marker='*', markersize=max(ms * 1.8, 11.0),
            color='#222222', linestyle='none', label='HV reference point',
            zorder=9)


def build_scatter_figure(exp, scn, styles, opts):
    """Scatter/front view for one scenario.

    opts keys: swap(bool), title(str|''), show_legend, show_labels,
    show_clouds, show_fronts, normalize(bool), hidden(set), marker_size(float),
    show_universal(bool, default True), show_ref_point(bool — mark the
    Hypervolume reference point of the current frame).

    PowerCap per-cap view: opts['tier'] scopes everything to that tier (base
    algorithm labels, the tier's universal front in black, normalization by the
    TIER's ideal/nadir) and opts['show_global_universal'] overlays the stored
    all-arms global universal front as a grey dashed reference.
    """
    fig = Figure(figsize=(9.2, 6.6), dpi=100)
    ax = fig.add_subplot(111)
    ox, oy = scn.obj_names
    if opts.get('swap'):
        ox, oy = oy, ox
    ms = float(opts.get('marker_size', 7))

    tier = opts.get('tier') if getattr(exp, 'is_powercap', False) else None
    if tier is not None:
        td = scn.tiers[tier]
        algos = [a for a in exp.base_algorithms if a not in opts.get('hidden', set())]
        points_src, fronts_src, universal = td.points, td.fronts, td.universal
        universal_label = f'Universal — {tier_display(tier)}'
        default_title = (f'{scn.name}: {axis_label(ox)} vs {axis_label(oy)} — '
                         f'{tier_display(tier, td.cap_watts)}')
        norm_bounds = tier_bounds(td, scn.obj_names)
    else:
        algos = _visible(exp, opts)
        points_src, fronts_src, universal = scn.points, scn.fronts, scn.universal
        universal_label = UNIVERSAL_LABEL
        default_title = f'{scn.name}: {axis_label(ox)} vs {axis_label(oy)}'
        norm_bounds = scenario_bounds(scn)

    if opts.get('normalize'):
        def val(series, col):
            ideal, span = norm_bounds[col]
            return (series - ideal) / span
    else:
        def val(series, col):
            return series

    for algo in algos:
        st = styles[algo]
        mfc = st['color'] if st['filled'] else 'none'
        if opts.get('show_clouds'):
            pts = points_src[points_src['Algorithm'] == algo]
            if len(pts):
                # gid drives the GUI's click-to-details lookup; the cloud is
                # plotted in file order so point index i within this algorithm
                # equals (seed, solutionIndex) row order of the CSV.
                coll = ax.scatter(val(pts[ox], ox), val(pts[oy], oy),
                                  s=(ms * 0.55) ** 2, alpha=0.22,
                                  color=st['color'], marker=st['marker'],
                                  linewidths=0, zorder=2)
                coll.set_gid(f'cloud::{algo}')
        if opts.get('show_fronts', True) and algo in fronts_src:
            fr = fronts_src[algo].sort_values(ox)
            line, = ax.plot(val(fr[ox], ox), val(fr[oy], oy),
                            color=st['color'], marker=st['marker'],
                            markersize=ms, markerfacecolor=mfc,
                            markeredgecolor=st['color'], linewidth=1.4, alpha=0.9,
                            label=st['display'], zorder=4)
            line.set_gid(f'front::{algo}')
            if opts.get('show_labels') and len(fr):
                mid = fr.iloc[len(fr) // 2]
                ax.annotate(st['display'],
                            (float(val(mid[ox], ox)), float(val(mid[oy], oy))),
                            textcoords='offset points', xytext=(6, 6),
                            fontsize=8, color=st['color'], fontweight='bold',
                            zorder=6)

    if (opts.get('show_universal', True)
            and universal is not None and len(universal)):
        uni = universal.sort_values(ox)
        ax.plot(val(uni[ox], ox), val(uni[oy], oy),
                color=UNIVERSAL_COLOR, linewidth=1.9,
                marker='o', markersize=max(ms * 0.5, 3.5),
                label=universal_label, zorder=8)

    if (tier is not None and opts.get('show_global_universal')
            and scn.universal is not None and len(scn.universal)):
        uni = scn.universal.sort_values(ox)
        ax.plot(val(uni[ox], ox), val(uni[oy], oy),
                color=GLOBAL_UNIVERSAL_COLOR, linewidth=1.6, linestyle='--',
                marker='o', markersize=max(ms * 0.4, 3.0),
                label='Global universal (all arms)', zorder=7)

    if opts.get('show_ref_point'):
        _draw_ref_point(ax, norm_bounds, ox, oy, opts.get('normalize'), ms)

    suffix = '  (normalized)' if opts.get('normalize') else ''
    ax.set_xlabel(axis_label(ox) + suffix)
    ax.set_ylabel(axis_label(oy) + suffix)
    ax.set_title(opts.get('title') or default_title)
    _apply_legend(ax, opts)
    fig.subplots_adjust(left=0.09, right=0.97, top=0.93, bottom=0.09)
    return fig


def build_compare_figure(exp, scn, base, styles, opts):
    """PowerCap compare mode: one base algorithm's pooled front at every cap
    tier, in the tier identity colors (Uncapped dark green + dashed, then
    light green / yellow / red loose -> tight). Normalization (when on) uses
    the scenario-wide frame, so it is comparable across tiers."""
    fig = Figure(figsize=(9.2, 6.6), dpi=100)
    ax = fig.add_subplot(111)
    ox, oy = scn.obj_names
    if opts.get('swap'):
        ox, oy = oy, ox
    ms = float(opts.get('marker_size', 7))
    marker = styles.get(base, {}).get('marker', 'o')
    display = styles.get(base, {}).get('display', base)

    bounds = scenario_bounds(scn)
    if opts.get('normalize'):
        def val(series, col):
            ideal, span = bounds[col]
            return (series - ideal) / span
    else:
        def val(series, col):
            return series

    hidden = opts.get('tiers_hidden', set())
    for tier in exp.tier_names:
        if tier in hidden:
            continue
        td = scn.tiers.get(tier)
        fr = td.fronts.get(base) if td is not None else None
        if fr is None or not len(fr):
            continue
        fr = fr.sort_values(ox)
        color = tier_color(exp, tier)
        line, = ax.plot(val(fr[ox], ox), val(fr[oy], oy),
                        color=color, marker=marker, markersize=ms,
                        markerfacecolor=color, markeredgecolor=color,
                        linewidth=1.6, alpha=0.95,
                        linestyle='--' if tier == UNCAPPED_TIER else '-',
                        label=tier_display(tier, td.cap_watts), zorder=4)
        line.set_gid(f'cmpfront::{tier}')
        if opts.get('show_labels') and len(fr):
            mid = fr.iloc[len(fr) // 2]
            ax.annotate(tier_display(tier),
                        (float(val(mid[ox], ox)), float(val(mid[oy], oy))),
                        textcoords='offset points', xytext=(6, 6),
                        fontsize=8, color=color, fontweight='bold', zorder=6)

    if opts.get('show_ref_point'):
        # Run-wide frame — the one the cross-cap Hypervolume is scored in.
        _draw_ref_point(ax, bounds, ox, oy, opts.get('normalize'), ms)

    suffix = '  (normalized, scenario-wide frame)' if opts.get('normalize') else ''
    ax.set_xlabel(axis_label(ox) + suffix)
    ax.set_ylabel(axis_label(oy) + suffix)
    ax.set_title(opts.get('title') or f'{scn.name}: {display} across power caps')
    _apply_legend(ax, opts)
    fig.subplots_adjust(left=0.09, right=0.97, top=0.93, bottom=0.09)
    return fig


def build_feasibility_figure(exp, styles, opts):
    """PowerCap only: per scenario, each UNCAPPED arm's mean feasibility rate
    under every derived cap (mean ± std over seeds, from
    feasibility_summary.csv), with the calibration targets as dashed lines."""
    fe = exp.feasibility
    fig = Figure(figsize=(9.2, 6.0), dpi=100)
    if fe is None or not len(fe):
        ax = fig.add_subplot(111)
        ax.text(0.5, 0.5, 'feasibility_summary.csv not found in this folder.',
                ha='center', va='center', color='#666666')
        ax.set_axis_off()
        return fig

    uncapped_arms = [a for a in exp.base_algorithms]
    hidden = opts.get('hidden', set())
    uncapped_arms = [a for a in uncapped_arms if a not in hidden]
    fe = fe[fe['Algorithm'].map(lambda a: split_tier(a)[1] == UNCAPPED_TIER)]
    caps = sorted(pd.unique(fe['CapWatts'].dropna()), reverse=True)
    pc_tiers = [t for t in exp.tier_names if t != UNCAPPED_TIER]
    scen_nums = sorted(pd.unique(fe['Scenario']))

    fig = Figure(figsize=(4.6 * max(len(scen_nums), 1) + 1.4, 5.6), dpi=100)
    axes = (list(np.atleast_1d(fig.subplots(1, len(scen_nums))))
            if len(scen_nums) > 1 else [fig.add_subplot(111)])
    width = 0.8 / max(len(caps), 1)
    for ax, s in zip(axes, scen_nums):
        sub = fe[fe['Scenario'] == s]
        x = np.arange(len(uncapped_arms))
        for ci, cap in enumerate(caps):
            tier = pc_tiers[ci] if ci < len(pc_tiers) else None
            color = tier_color(exp, tier) if tier else FALLBACK_COLORS[ci]
            means, stds = [], []
            for algo in uncapped_arms:
                row = sub[(sub['Algorithm'] == algo) & (sub['CapWatts'] == cap)]
                means.append(float(row['MeanFeasibilityRate'].iloc[0]) if len(row) else np.nan)
                stds.append(float(row['StdFeasibilityRate'].iloc[0]) if len(row) else np.nan)
            label = (tier_display(tier, cap) if tier
                     else f'{cap / 1000.0:.1f} kW')
            ax.bar(x + ci * width - 0.4 + width / 2,
                   [0 if np.isnan(m) else m for m in means], width * 0.92,
                   yerr=[0 if np.isnan(sd) else sd for sd in stds],
                   capsize=2, error_kw={'linewidth': 0.8},
                   color=color, label=label)
        for tier in pc_tiers:
            ax.axhline(int(tier[2:]) / 100.0, color='#888888',
                       linestyle='--', linewidth=0.9, zorder=1)
        ax.set_xticks(x)
        ax.set_xticklabels([styles.get(a, {'display': a})['display']
                            for a in uncapped_arms],
                           rotation=38, ha='right', fontsize=8)
        name = ''
        named = sub.drop_duplicates('Scenario')
        if len(named) and 'ScenarioName' in named:
            name = str(named['ScenarioName'].iloc[0]).replace('_', ' ')
        ax.set_title(name or f'Scenario {s}', fontsize=11)
        ax.set_ylim(0, 1.05)
    axes[0].set_ylabel('Mean feasibility rate')
    if opts.get('show_legend', True):
        axes[-1].legend(fontsize=8, framealpha=0.85, loc='upper right')
    fig.suptitle(opts.get('title') or
                 'Uncapped arms — share of solutions feasible under each derived cap '
                 '(dashed: calibration targets)', fontsize=12)
    fig.subplots_adjust(left=0.07, right=0.985, top=0.86, bottom=0.28, wspace=0.24)
    return fig


# Mirrors plot_scenario_pareto.AXIS_LABEL (Energy is kWh in every CSV).
AXIS_UNITS = {
    'Makespan': 'Makespan (s)',
    'WaitingTime': 'Average Waiting Time (s)',
    'Energy': 'Energy (kWh)',
}


def axis_label(obj_name):
    return AXIS_UNITS.get(obj_name, obj_name)


def _metric_values(scn, algo, key):
    """(mean, std) over seeds of one per-seed metric for one algorithm."""
    if scn.metrics_seed is None:
        return (np.nan, np.nan)
    rows = scn.metrics_seed[scn.metrics_seed['Algorithm'] == algo]
    if key not in rows.columns or not len(rows):
        return (np.nan, np.nan)
    vals = rows[key].dropna()
    if not len(vals):
        return (np.nan, np.nan)
    return (vals.mean(), vals.std(ddof=1) if len(vals) > 1 else 0.0)


# =============================================================================
# COLLECTIVE (pooled-front) INDICATORS — the "10-run Pareto set" view
# =============================================================================

GLOBAL_TABLE_KIND = 'Global metrics — collective pooled fronts'
GLOBAL_COLUMNS = ['Hypervolume', 'GD', 'IGD', 'Spacing', 'Front size',
                  'Contribution', '% of universal', 'Avg runtime (s)']
# build_metric_figure key -> collective_metrics row key
COLLECTIVE_KEY = {'HV_fixed': 'Hypervolume', 'GD': 'GD', 'IGD': 'IGD',
                  'Spacing': 'Spacing', 'ParetoContribution': 'Contribution'}

# Tables-tab per-seed pivots: (display name, CSV column). One hypervolume, one
# contribution; the legacy HV / TotalSolutions columns are not offered.
PER_SEED_TABLES = [
    ('Hypervolume', 'HV_fixed'),
    ('GD', 'GD'),
    ('IGD', 'IGD'),
    ('Spacing', 'Spacing'),
    ('Front size', 'NonDomSolutions'),
    ('Pareto contribution', 'ParetoContribution'),
    ('Runtime (ms)', 'TimeMs'),
]
PER_SEED_TABLE_COLS = dict(PER_SEED_TABLES)

# PowerCap compare mode tables: display name -> compare_pivot_table metric.
COMPARE_TABLE_KINDS = {
    'Across caps — Hypervolume (run-wide frame)': 'HV_fixed_run',
    'Across caps — GD (tier reference)': 'GD',
    'Across caps — IGD (tier reference)': 'IGD',
    'Across caps — Spacing': 'Spacing',
}


def _src_union_bounds(src, obj_names):
    """Pooled ideal/nadir of one source (ScenarioData or TierData): all
    published points + the universal front — exactly the HV_fixed frame."""
    frames = [src.points[obj_names].to_numpy(dtype=float)]
    if src.universal is not None and len(src.universal):
        frames.append(src.universal[obj_names].to_numpy(dtype=float))
    return _union_bounds(np.vstack(frames))


def hv_frame_info(bounds, obj_names):
    """Human-readable description of the Hypervolume frame + reference point."""
    ix, iy, nx, ny = bounds
    rx = nx + 0.1 * (nx - ix)
    ry = ny + 0.1 * (ny - iy)
    return (f'HV frame — {obj_names[0]}: ideal {ix:.6g}, nadir {nx:.6g} · '
            f'{obj_names[1]}: ideal {iy:.6g}, nadir {ny:.6g} · '
            f'reference point ({rx:.6g}, {ry:.6g}) = nadir + 10% of range '
            '(1.1, 1.1 normalized)')


def collective_metrics(src, obj_names, algorithms):
    """Indicators of each algorithm's COLLECTIVE Pareto set — the pooled
    union-over-seeds front — all in the source's shared frame:

      Hypervolume   HV_fixed of the pooled front (fixed reference point)
      GD / IGD / Spacing   pooled front vs the universal front (same math as
                           the per-seed rows, applied to the union front)
      Front size    points on the pooled front
      Contribution  strict union-over-seeds count (the CSV MEAN row when
                    present, else recomputed from the point clouds)
      % of universal, Avg runtime (s)

    Works on a ScenarioData or a TierData (duck-typed). Returns
    ({algo: {column: value}}, universal_row, bounds)."""
    if src.points is None or not len(src.points):
        return {}, {}, None
    bounds = _src_union_bounds(src, obj_names)
    uni = (src.universal[obj_names].to_numpy(dtype=float)
           if src.universal is not None and len(src.universal)
           else np.empty((0, 2)))
    out = {}
    for algo in algorithms:
        cloud = src.points[src.points['Algorithm'] == algo]
        fr = src.fronts.get(algo)
        pooled = (fr[list(obj_names)].to_numpy(dtype=float)
                  if fr is not None and len(fr)
                  else cloud[obj_names].to_numpy(dtype=float))
        if not len(pooled):
            continue
        nd = _filter_nondominated(pooled)
        _, gd, igd, spacing = _legacy_pair_metrics(nd, uni)
        if not len(uni):
            gd = igd = np.nan
        contrib = np.nan
        if (src.metrics_mean is not None and algo in src.metrics_mean.index
                and 'ParetoContribution' in src.metrics_mean.columns):
            contrib = src.metrics_mean.loc[algo, 'ParetoContribution']
        if pd.isna(contrib) and len(uni) and len(cloud):
            runs = [g[obj_names].to_numpy(dtype=float)
                    for _, g in cloud.groupby('Seed', sort=False)]
            contrib = _union_contribution_count(runs, uni)
        time_s = np.nan
        if src.metrics_seed is not None and 'TimeMs' in src.metrics_seed.columns:
            times = src.metrics_seed.loc[
                src.metrics_seed['Algorithm'] == algo, 'TimeMs'].dropna()
            times = times[times > 0]
            if len(times):
                time_s = float(times.mean()) / 1000.0
        out[algo] = {
            'Hypervolume': _hv_fixed(_normalize_union(pooled, bounds)),
            'GD': gd, 'IGD': igd, 'Spacing': spacing,
            'Front size': int(len(nd)),
            'Contribution': float(contrib) if pd.notna(contrib) else np.nan,
            '% of universal': (100.0 * contrib / len(uni)
                               if pd.notna(contrib) and len(uni) else np.nan),
            'Avg runtime (s)': time_s,
        }
    universal_row = {}
    if len(uni):
        universal_row = {
            'Hypervolume': _hv_fixed(_normalize_union(uni, bounds)),
            'GD': 0.0, 'IGD': 0.0, 'Spacing': np.nan,
            'Front size': int(len(uni)), 'Contribution': float(len(uni)),
            '% of universal': 100.0, 'Avg runtime (s)': np.nan,
        }
    return out, universal_row, bounds


def global_metric_table(src, obj_names, algorithms):
    """The Global metrics table: one row per algorithm (+ the universal front),
    indicators of the collective pooled fronts. Returns (DataFrame indexed by
    Algorithm, info string) or (None, '')."""
    rows, universal_row, bounds = collective_metrics(src, obj_names, algorithms)
    if not rows:
        return None, ''
    df = pd.DataFrame.from_dict(rows, orient='index').reindex(columns=GLOBAL_COLUMNS)
    if universal_row:
        df.loc[UNIVERSAL_KEY] = pd.Series(universal_row)
    df.index.name = 'Algorithm'
    n_seeds = ''
    if src.metrics_seed is not None and len(src.metrics_seed):
        n_seeds = f"{src.metrics_seed['Seed'].nunique()} "
    info = (f'Collective = each algorithm\'s pooled front (union of the '
            f'{n_seeds}seed runs, non-dominated). Hypervolume uses the fixed '
            'reference point in the pooled frame; GD/IGD are measured against '
            f'the universal front ({int(universal_row.get("Front size", 0))} '
            'points); Contribution = strict matches on the universal front.')
    return df, info


def build_metric_figure(exp, key, display, ylabel, source, styles, opts):
    """One indicator, one panel per scenario, bars = algorithms.

    opts['aggregation'] picks the number behind each bar: 'collective'
    (default) scores each algorithm's pooled union-over-seeds front (single
    bars, no error bar; Hypervolume additionally draws the universal front's
    value as a dashed ceiling line), 'mean' is the per-seed mean ± std.

    PowerCap: opts['tier'] scopes bars to the tier's arms with tier-referenced
    values; compare mode (opts['mode'] == 'cmp') plots one base arm with one
    bar per cap tier instead — `source` is only used there (see
    COMPARE_METRIC_CHOICES; pass None otherwise)."""
    scns = [exp.scenarios[n] for n in sorted(exp.scenarios)]
    fig = Figure(figsize=(4.6 * max(len(scns), 1) + 1.2, 5.4), dpi=100)
    axes = fig.subplots(1, len(scns)) if len(scns) > 1 else [fig.add_subplot(111)]
    if len(scns) > 1:
        axes = list(np.atleast_1d(axes))

    powercap = getattr(exp, 'is_powercap', False)
    compare = powercap and opts.get('mode') == 'cmp'
    tier = opts.get('tier') if powercap else None

    if compare:
        base = opts.get('cmp_algo')
        hidden = opts.get('tiers_hidden', set())
        tiers = [t for t in exp.tier_names if t not in hidden]
        for ax, scn in zip(axes, scns):
            means, stds, colors, names = [], [], [], []
            for t in tiers:
                m, s = _compare_metric_values(exp, scn, base, t, key, source)
                means.append(m)
                stds.append(s)
                colors.append(tier_color(exp, t))
                names.append(tier_display(t, scn.tiers.get(t).cap_watts
                                          if scn.tiers.get(t) else None))
            x = np.arange(len(tiers))
            yerr = None if all(np.isnan(s) for s in stds) else \
                [0 if np.isnan(s) else s for s in stds]
            ax.bar(x, [0 if np.isnan(m) else m for m in means], color=colors,
                   yerr=yerr, capsize=3, error_kw={'linewidth': 1})
            ax.set_xticks(x)
            ax.set_xticklabels(names, rotation=38, ha='right', fontsize=8)
            ax.set_title(scn.name, fontsize=11)
            ax.grid(axis='x', alpha=0)
        axes[0].set_ylabel(ylabel)
        disp = styles.get(base, {}).get('display', base)
        fig.suptitle(opts.get('title') or f'{disp} — {display}', fontsize=13)
        fig.subplots_adjust(left=0.07, right=0.985, top=0.86, bottom=0.28,
                            wspace=0.24)
        return fig

    aggregation = opts.get('aggregation', 'collective')
    algos = ([a for a in exp.base_algorithms if a not in opts.get('hidden', set())]
             if tier is not None else _visible(exp, opts))
    for ax, scn in zip(axes, scns):
        src = scn.tiers.get(tier) if tier is not None else scn
        means, stds, colors, names = [], [], [], []
        corner = ''
        universal_hv = None
        if aggregation == 'collective':
            cm, universal_row, _ = collective_metrics(src, scn.obj_names, algos)
            col = COLLECTIVE_KEY.get(key, key)
            for algo in algos:
                means.append(cm.get(algo, {}).get(col, np.nan))
                stds.append(np.nan)
                colors.append(styles[algo]['color'])
                names.append(styles[algo]['display'])
            if universal_row:
                if key == 'HV_fixed':
                    universal_hv = universal_row.get('Hypervolume')
                    corner = f'universal front: {universal_hv:.3f}'
                elif key == 'ParetoContribution':
                    corner = f'universal front: {int(universal_row["Front size"])} pts'
        else:
            for algo in algos:
                m, s = (_tier_metric_values(src, algo, key, 'seed')
                        if tier is not None else _metric_values(scn, algo, key))
                means.append(m)
                stds.append(s)
                colors.append(styles[algo]['color'])
                names.append(styles[algo]['display'])
            if key == 'ParetoContribution':
                uni = src.universal
                if uni is not None and len(uni):
                    corner = f'universal front: {len(uni)} pts'
        x = np.arange(len(algos))
        yerr = None if all(np.isnan(s) for s in stds) else \
            [0 if np.isnan(s) else s for s in stds]
        ax.bar(x, [0 if np.isnan(m) else m for m in means], color=colors,
               yerr=yerr, capsize=3, error_kw={'linewidth': 1})
        if universal_hv is not None and np.isfinite(universal_hv):
            ax.axhline(universal_hv, color='#333333', linestyle='--',
                       linewidth=1.2, zorder=1)
        ax.set_xticks(x)
        ax.set_xticklabels(names, rotation=38, ha='right', fontsize=8)
        ax.set_title(scn.name, fontsize=11)
        ax.grid(axis='x', alpha=0)
        if corner:
            ax.text(0.98, 0.96, corner, transform=ax.transAxes, ha='right',
                    va='top', fontsize=8, color='#555555')
    axes[0].set_ylabel(ylabel)
    title = opts.get('title') or (
        f'{display} — collective pooled fronts (union over seeds)'
        if aggregation == 'collective' else f'{display} — mean ± std over seeds')
    if tier is not None and not opts.get('title'):
        title += f' — {tier_display(tier, exp.cap_watts.get(tier))}'
    fig.suptitle(title, fontsize=13)
    fig.subplots_adjust(left=0.07, right=0.985, top=0.86, bottom=0.28,
                        wspace=0.24)
    return fig


def build_runtime_figure(exp, styles, opts):
    """Average runtime (s) per algorithm: one panel per scenario + overall.

    PowerCap per-cap view scopes to the tier's runs; compare mode shows one
    base arm's runtime with one bar per tier."""
    scns = [exp.scenarios[n] for n in sorted(exp.scenarios)]
    powercap = getattr(exp, 'is_powercap', False)
    compare = powercap and opts.get('mode') == 'cmp'
    tier = opts.get('tier') if powercap else None

    if compare:
        base = opts.get('cmp_algo')
        hidden = opts.get('tiers_hidden', set())
        tiers = [t for t in exp.tier_names if t not in hidden]
        n = max(len(scns), 1)
        fig = Figure(figsize=(4.6 * n + 1.2, 5.4), dpi=100)
        axes = (list(np.atleast_1d(fig.subplots(1, n)))
                if n > 1 else [fig.add_subplot(111)])
        for ax, scn in zip(axes, scns):
            means, stds, colors, names = [], [], [], []
            for t in tiers:
                td = scn.tiers.get(t)
                rows = (td.metrics_seed if td is not None
                        and td.metrics_seed is not None else None)
                vals = (rows[rows['Algorithm'] == base]['TimeMs'].dropna() / 1000.0
                        if rows is not None else pd.Series(dtype=float))
                means.append(vals.mean() if len(vals) else 0)
                stds.append(vals.std(ddof=1) if len(vals) > 1 else 0)
                colors.append(tier_color(exp, t))
                names.append(tier_display(t))
            x = np.arange(len(tiers))
            ax.bar(x, means, yerr=stds, color=colors, capsize=3,
                   error_kw={'linewidth': 1})
            ax.set_xticks(x)
            ax.set_xticklabels(names, rotation=38, ha='right', fontsize=8)
            ax.set_title(scn.name, fontsize=11)
        axes[0].set_ylabel('runtime (s)')
        disp = styles.get(base, {}).get('display', base)
        fig.suptitle(opts.get('title')
                     or f'{disp} — average runtime across power caps', fontsize=13)
        fig.subplots_adjust(left=0.07, right=0.985, top=0.86, bottom=0.28,
                            wspace=0.24)
        return fig

    if tier is not None:
        panels = [(s.name, s.tiers[tier].metrics_seed) for s in scns
                  if tier in s.tiers and s.tiers[tier].metrics_seed is not None]
        algos = [a for a in exp.base_algorithms if a not in opts.get('hidden', set())]
        subtitle = f' — {tier_display(tier, exp.cap_watts.get(tier))}'
    else:
        panels = [(s.name, s.metrics_seed) for s in scns if s.metrics_seed is not None]
        algos = _visible(exp, opts)
        subtitle = ''
    all_rows = pd.concat([p[1] for p in panels], ignore_index=True) if panels else None
    if all_rows is not None and len(panels) > 1:
        panels = panels + [('All scenarios', all_rows)]
    n = max(len(panels), 1)
    fig = Figure(figsize=(4.6 * n + 1.2, 5.4), dpi=100)
    axes = list(np.atleast_1d(fig.subplots(1, n))) if n > 1 else [fig.add_subplot(111)]

    for ax, (name, rows) in zip(axes, panels):
        means, stds, colors, names = [], [], [], []
        for algo in algos:
            vals = rows[rows['Algorithm'] == algo]['TimeMs'].dropna() / 1000.0
            means.append(vals.mean() if len(vals) else 0)
            stds.append(vals.std(ddof=1) if len(vals) > 1 else 0)
            colors.append(styles[algo]['color'])
            names.append(styles[algo]['display'])
        x = np.arange(len(algos))
        ax.bar(x, means, yerr=stds, color=colors, capsize=3,
               error_kw={'linewidth': 1})
        ax.set_xticks(x)
        ax.set_xticklabels(names, rotation=38, ha='right', fontsize=8)
        ax.set_title(name, fontsize=11)
    if panels:
        axes[0].set_ylabel('runtime (s)')
    fig.suptitle((opts.get('title')
                  or 'Average runtime per algorithm (mean ± std over seeds)' + subtitle),
                 fontsize=13)
    fig.subplots_adjust(left=0.07, right=0.985, top=0.86, bottom=0.28,
                        wspace=0.24)
    return fig


# =============================================================================
# TABLE BUILDERS
# =============================================================================

def pivot_metric_table(scn, metric):
    """(DataFrame seeds×algorithms + MEAN/STDDEV rows, universal-info string)."""
    if scn.metrics_seed is None or metric not in scn.metrics_seed.columns:
        return None, ''
    piv = scn.metrics_seed.pivot_table(index='Seed', columns='Algorithm',
                                       values=metric, sort=False)
    piv = piv[[a for a in scn.metrics_seed['Algorithm'].unique() if a in piv.columns]]
    piv.index = piv.index.astype(str)
    extra = {}
    if scn.metrics_mean is not None and metric in scn.metrics_mean.columns:
        extra['MEAN'] = {a: scn.metrics_mean.loc[a, metric]
                         for a in piv.columns if a in scn.metrics_mean.index}
    if scn.metrics_std is not None and metric in scn.metrics_std.columns:
        extra['STDDEV'] = {a: scn.metrics_std.loc[a, metric]
                           for a in piv.columns if a in scn.metrics_std.index}
    for name, row in extra.items():
        piv.loc[name] = pd.Series(row)

    info = ''
    um = scn.universal_metrics
    if um:
        bits = [f"size={int(um['NonDomSolutions'])}"
                if pd.notna(um.get('NonDomSolutions')) else None,
                f"Hypervolume={um['HV_fixed']:.6f}"
                if pd.notna(um.get('HV_fixed')) else None]
        info = 'Pooled universal front: ' + ', '.join(b for b in bits if b)
    return piv, info


def compare_pivot_table(exp, scn, base, metric):
    """Compare-mode table: seeds × cap tiers for one base arm.

    metric 'HV_fixed_run' reads the stored run-wide-frame HV_fixed rows of the
    tier-suffixed arm from the GLOBAL metrics (comparable across caps); any
    other metric reads the per-tier tables (referenced to each tier's own
    universal front)."""
    cols = {}
    for tier in exp.tier_names:
        td = scn.tiers.get(tier)
        name = tier_display(tier)
        if metric == 'HV_fixed_run':
            if scn.metrics_seed is None or 'HV_fixed' not in scn.metrics_seed.columns:
                continue
            full = base if tier == UNCAPPED_TIER else f'{base}_{tier}'
            rows = scn.metrics_seed[scn.metrics_seed['Algorithm'] == full]
            if len(rows):
                cols[name] = rows.set_index('Seed')['HV_fixed']
        else:
            if td is None or td.metrics_seed is None or metric not in td.metrics_seed.columns:
                continue
            rows = td.metrics_seed[td.metrics_seed['Algorithm'] == base]
            if len(rows):
                cols[name] = rows.set_index('Seed')[metric]
    if not cols:
        return None, ''
    piv = pd.DataFrame(cols)
    piv.index = piv.index.astype(str)
    means = piv.mean()
    stds = piv.std(ddof=0)
    piv.loc['MEAN'] = means
    piv.loc['STDDEV'] = stds
    info = ('Hypervolume (run-wide frame) is comparable across caps; '
            'GD/IGD/Spacing reference each tier\'s own universal front.')
    return piv, info


def collab_table(scn):
    """Seed-collaboration scoreboard as a display frame (or None)."""
    if scn.collab_seed is None:
        return None, ''
    piv = scn.collab_seed.pivot_table(index='Seed', columns='Algorithm',
                                      values='ContributionPct', sort=False)
    piv = piv[[a for a in scn.collab_seed['Algorithm'].unique() if a in piv.columns]]
    piv.index = piv.index.astype(str)
    if scn.collab_mean is not None:
        piv.loc['MEAN'] = pd.Series({a: scn.collab_mean.loc[a, 'ContributionPct']
                                     for a in piv.columns if a in scn.collab_mean.index})
    if scn.collab_std is not None:
        piv.loc['STDDEV'] = pd.Series({a: scn.collab_std.loc[a, 'ContributionPct']
                                       for a in piv.columns if a in scn.collab_std.index})
    info = ''
    cu = scn.collab_universal
    if 'mean' in cu:
        mu, sd = cu['mean'], cu.get('std', {})
        info = ('Per-seed universal front: size '
                f"{mu.get('SeedUniversalFrontSize', float('nan')):.1f}"
                f" ± {sd.get('SeedUniversalFrontSize', float('nan')):.1f}, "
                f"HV_fixed {mu.get('SeedUniversalHV_fixed', float('nan')):.4f}"
                f" ± {sd.get('SeedUniversalHV_fixed', float('nan')):.4f}"
                ' (near-tie shares can sum >100%)')
    return piv, info


# =============================================================================
# DETAILS (scenario_N_solution_details.json) — helpers
# =============================================================================

def details_solutions(scn):
    if not scn.details:
        return []
    return scn.details.get('solutions', [])


def details_summary_rows(sol, scn):
    """Aggregate figures of one solution as (name, value-string) pairs."""
    rows = []
    obj_names = (scn.details or {}).get('objectives', scn.obj_names)
    for name, val in zip(obj_names, sol.get('objectives') or []):
        rows.append((f'Objective: {axis_label(name)}', f'{val:.6f}'))
    for key, label, fmt in [
            ('makespanSeconds', 'Makespan (s)', '{:.4f}'),
            ('avgWaitSeconds', 'Avg wait time (s)', '{:.4f}'),
            ('totalEnergyKWh', 'Total energy (kWh)', '{:.6f}'),
            ('peakPowerWatts', 'Peak total power (W)', '{:.2f}'),
            ('avgQueueSizeAllVms', 'Avg queue size (all VMs)', '{:.2f}'),
            ('avgQueueSizeActiveVms', 'Avg queue size (active VMs)', '{:.2f}')]:
        if sol.get(key) is not None:
            rows.append((label, fmt.format(sol[key])))
    queues = sol.get('vmQueues')
    if queues is not None:
        active = sol.get('activeVmCount', sum(1 for q in queues if q))
        rows.append(('Active VMs', f'{active} / {len(queues)}'))
    if sol.get('roles'):
        rows.append(('Representative roles', ', '.join(sol['roles'])))
    return rows


def details_vm_table(sol, scn):
    """Per-VM queue table rows from a solution dict."""
    vms = {v['index']: v for v in (scn.details.get('vms') or [])}
    rows = []
    for idx, queue in enumerate(sol.get('vmQueues') or []):
        meta = vms.get(idx, {})
        preview = ','.join(str(t) for t in queue[:12]) + (',…' if len(queue) > 12 else '')
        rows.append((idx, meta.get('computeType', ''), meta.get('vcpus', ''),
                     meta.get('gpus', ''), meta.get('hostIndex', ''),
                     len(queue), preview))
    return rows


def details_host_table(sol, scn):
    hosts = {h['index']: h for h in (scn.details.get('hosts') or [])}
    rows = []
    for idx, kwh in enumerate(sol.get('hostEnergyKWh') or []):
        meta = hosts.get(idx, {})
        rows.append((idx, meta.get('computeType', ''), meta.get('cores', ''),
                     meta.get('gpus', ''), f'{kwh:.9f}'))
    return rows


def details_task_table(sol, scn):
    tasks = {t['index']: t for t in (scn.details.get('tasks') or [])}
    assign = sol.get('taskVmIndex') or []
    waits = sol.get('taskWaitSeconds') or []
    rows = []
    for tid, vm in enumerate(assign):
        meta = tasks.get(tid, {})
        wait = f'{waits[tid]:.3f}' if tid < len(waits) else ''
        rows.append((tid, meta.get('lengthInstructions', ''),
                     meta.get('workload', ''), vm, wait))
    return rows


def solution_label(sol, obj_names):
    """Human-readable identity of one solution for pickers/status lines."""
    objs = sol.get('objectives') or []
    parts = [f'#{sol.get("solutionIndex", "?")}']
    parts += [f'{n}={v:.6g}' for n, v in zip(obj_names, objs)]
    if sol.get('roles'):
        parts.append('[' + ', '.join(sol['roles']) + ']')
    return '  '.join(parts)


# =============================================================================
# SAVE FOR CLAUDE — one self-contained Markdown bundle for LLM analysis
# =============================================================================

def _df_to_csv_block(df, index_label=None, float_format='%.6f'):
    txt = df.to_csv(index_label=index_label, float_format=float_format)
    return '```csv\n' + txt + '```\n'


def _bundle_indicator_sections(L, src, where, obj_names, universal_name):
    """Emits the indicator/collaboration/front sections for one source — a
    ScenarioData or (duck-typed) TierData. `where` names it in headings."""
    if src.metrics_seed is not None:
        L.append(f'### Per-seed quality indicators ({where})')
        L.append(_df_to_csv_block(src.metrics_seed, float_format='%.6g'))
        mixed = []
        if src.metrics_mean is not None:
            mm = src.metrics_mean.copy()
            mm.insert(0, 'Kind', 'MEAN')
            mixed.append(mm)
        if src.metrics_std is not None:
            ms = src.metrics_std.copy()
            ms.insert(0, 'Kind', 'STDDEV')
            mixed.append(ms)
        if mixed:
            L.append(f'### MEAN / STDDEV rows ({where})')
            L.append(_df_to_csv_block(pd.concat(mixed).drop(columns=['Seed'],
                                                            errors='ignore'),
                                      index_label='Algorithm',
                                      float_format='%.6g'))
        if src.universal_metrics:
            um = {k: v for k, v in src.universal_metrics.items()
                  if k not in ('Algorithm', 'Seed') and pd.notna(v)}
            L.append(f'{universal_name} metrics: '
                     + ', '.join(f'{k}={v:.6g}' if isinstance(v, float)
                                 else f'{k}={v}' for k, v in um.items()))
            L.append('')
    if src.collab_seed is not None:
        L.append(f'### Seed-collaboration scoreboard (near-tie credit, {where})')
        L.append(_df_to_csv_block(src.collab_seed, float_format='%.6g'))
        piv, info = collab_table(src)
        if info:
            L.append(info)
            L.append('')
    if src.universal is not None and len(src.universal):
        L.append(f'### {universal_name} ({len(src.universal)} points)')
        L.append(_df_to_csv_block(src.universal[obj_names], float_format='%.9g'))
    if src.fronts:
        L.append(f'### Per-algorithm pooled fronts (union over seeds, {where})')
        for label, fr in src.fronts.items():
            L.append(f'#### {label} ({len(fr)} points)')
            L.append(_df_to_csv_block(fr[obj_names], float_format='%.9g'))


def build_claude_bundle(exp):
    L = []
    objs = exp.objective_names
    L.append(f'# Experiment bundle: {exp.name}')
    L.append('')
    L.append(f'- Generated: {datetime.now(timezone.utc).isoformat(timespec="seconds")}')
    L.append(f'- Study objectives (both minimized): {objs[0]} vs {objs[1]}')
    L.append(f'- Scenarios: ' + ', '.join(
        f'{n} = {exp.scenarios[n].name}' for n in sorted(exp.scenarios)))
    if exp.is_powercap:
        L.append(f'- Base algorithms: ' + ', '.join(exp.base_algorithms))
        L.append('- Cap tiers: ' + ', '.join(
            tier_display(t, exp.cap_watts.get(t)) for t in exp.tier_names))
        L.append(f'- Per-tier data source: {exp.by_cap_source}')
    else:
        L.append(f'- Algorithms: ' + ', '.join(exp.algorithms))
    L.append('')
    L.append('Notes for analysis: HV_fixed (when present) is the only hypervolume '
             'comparable across algorithms/seeds within this run; the legacy HV '
             'column uses a per-pair normalization frame and is NOT cross-comparable. '
             'ParetoContribution per-seed rows are strict exact-match counts against '
             'the pooled universal front; the MEAN row of that column is the '
             'union-over-seeds count. The seed-collaboration table uses near-tie '
             'credit (0.3% relative on both objectives), so shares can sum to >100%.')
    if exp.is_powercap:
        L.append('')
        L.append('PowerCeiling study: this bundle is organized per CAP TIER. Within a '
                 'tier section, every indicator (HV_fixed, GD, IGD, Spacing, '
                 'contribution, collaboration) is computed strictly against that '
                 'tier\'s own universal front and normalization bounds — the tier '
                 'frame. Tier-frame HV_fixed values are NOT comparable across tiers; '
                 'for cross-cap comparisons of one arm use the run-wide-frame '
                 'HV_fixed in the "global (all arms)" appendix, whose frame pools '
                 'all tiers. The appendix\'s global universal front mixes capped and '
                 'uncapped regimes — treat it as an envelope, not a per-cap reference.')
    L.append('')

    for n in sorted(exp.scenarios):
        scn = exp.scenarios[n]
        L.append(f'## Scenario {n}: {scn.name}')
        L.append('')
        if exp.is_powercap:
            for tier in exp.tier_names:
                td = scn.tiers.get(tier)
                if td is None:
                    continue
                L.append(f'## Scenario {n} — {tier_display(tier, td.cap_watts)}')
                L.append('')
                _bundle_indicator_sections(
                    L, td, f'scenario {n}, {tier_display(tier)}', scn.obj_names,
                    f'{tier_display(tier)} universal Pareto front')
            L.append(f'## Scenario {n} — appendix: global (all arms, mixed regimes)')
            L.append('')
            if scn.universal is not None and len(scn.universal):
                L.append(f'### Global universal front ({len(scn.universal)} points, '
                         'pooled over uncapped + all caps)')
                L.append(_df_to_csv_block(scn.universal[scn.obj_names],
                                          float_format='%.9g'))
            if scn.metrics_seed is not None and 'HV_fixed' in scn.metrics_seed.columns:
                L.append('### Run-wide-frame HV_fixed per (arm, tier, seed) — '
                         'comparable across caps')
                run_frame = scn.metrics_seed[['Algorithm', 'Seed', 'HV_fixed']]
                L.append(_df_to_csv_block(run_frame, float_format='%.6g'))
        else:
            _bundle_indicator_sections(L, scn, f'scenario {n}', scn.obj_names,
                                       'Pooled universal Pareto front')
        sols = details_solutions(scn)
        if sols:
            reps = [s for s in sols if s.get('roles')] or sols
            L.append(f'### Solution schedule details (scenario {n}) — '
                     f'{len(sols)} front solutions captured, aggregates below for '
                     f'the {len(reps)} representative (best-per-objective / knee) ones')
            L.append('Full per-solution task→VM maps, per-VM queues and per-host '
                     'energy are in scenario_N_solution_details.json.gz in the '
                     'result folder.')
            L.append('')
            obj_names = (scn.details or {}).get('objectives', scn.obj_names)
            rows = []
            for sol in reps:
                row = {'algorithm': sol.get('algorithm'),
                       'seed': sol.get('seed'),
                       'roles': '+'.join(sol.get('roles') or [])}
                row.update({f'obj_{k}': v for k, v in
                            zip(obj_names, sol.get('objectives') or [])})
                for k in ('makespanSeconds', 'avgWaitSeconds', 'totalEnergyKWh',
                          'peakPowerWatts', 'avgQueueSizeAllVms', 'activeVmCount'):
                    if sol.get(k) is not None:
                        row[k] = sol[k]
                rows.append(row)
            L.append(_df_to_csv_block(pd.DataFrame(rows), float_format='%.6g'))
        L.append('')
    if exp.is_powercap and exp.feasibility is not None and len(exp.feasibility):
        L.append('## Feasibility summary (every arm × derived cap, from '
                 'feasibility_summary.csv)')
        L.append(_df_to_csv_block(exp.feasibility, float_format='%.6g'))
    return '\n'.join(L)


# =============================================================================
# INSTRUCTIONS_FOR_CLAUDE.md — generated analysis briefing for the zip bundle
# =============================================================================

def build_instructions_md(exp, file_map):
    """A read-first briefing for the Claude instance that analyzes the bundle:
    what this run is, what every file contains, how to read the numbers, and
    what analysis to perform. Generated from the loaded folder, not boilerplate.

    file_map: list of (relative path, one-line description) for the zip contents.
    """
    objs = exp.objective_names
    L = [f'# How to analyze this bundle — {exp.name}', '']
    L.append('You are looking at the complete result bundle of a cloud '
             'task-scheduling experiment from the COSMOS simulator. Read this '
             'file fully before touching the data.')
    L.append('')
    L.append('## What this run is')
    L.append('')
    L.append(f'- Study: {objs[0]} vs {objs[1]} — BOTH objectives are minimized.')
    L.append(f'- Scenarios: ' + '; '.join(
        f'{n} = {exp.scenarios[n].name}' for n in sorted(exp.scenarios)))
    first = next(iter(exp.scenarios.values()))
    n_seeds = (first.metrics_seed['Seed'].nunique()
               if first.metrics_seed is not None else 0)
    if exp.is_powercap:
        L.append(f'- PowerCeiling study: {len(exp.base_algorithms)} base arms '
                 f'({", ".join(exp.base_algorithms)}), each run uncapped and '
                 f're-run under every derived power cap.')
        L.append('- Cap tiers: ' + ', '.join(
            tier_display(t, exp.cap_watts.get(t)) for t in exp.tier_names)
            + ' (tier = target feasibility percent; caps were calibrated from '
              'the uncapped peak-power distribution).')
        L.append(f'- Per-tier tables were '
                 + ('written natively by the Java analyzer (*_by_cap.csv).'
                    if exp.by_cap_source == 'native' else
                    'recomputed by the explorer from the point clouds '
                    '(pre-fix folder without *_by_cap.csv).'))
    else:
        L.append(f'- Algorithms: {", ".join(exp.algorithms)}.')
    L.append(f'- Seeds per (arm, scenario): {n_seeds}. Every metric row is one '
             '(arm, seed) run; MEAN/STDDEV rows aggregate over seeds.')
    L.append('')
    L.append('## Files in this bundle')
    L.append('')
    for rel, desc in file_map:
        L.append(f'- `{rel}` — {desc}')
    L.append('')
    L.append('## How to read the numbers (important)')
    L.append('')
    L.append('- `HV_fixed` is the only hypervolume comparable across arms and '
             'seeds; higher is better, range [0, 1]. The legacy `HV` column uses '
             'a per-pair normalization frame and must NOT be compared across '
             'arms — do not cite it.')
    L.append('- `GD`/`IGD` are distances to the universal (reference) front '
             '(lower is better); `Spacing` measures only front regularity, not '
             'quality.')
    L.append('- `ParetoContribution` per-seed rows are strict 1e-9 exact-match '
             'counts against the pooled universal front; the MEAN row of that '
             'column is the union-over-seeds count, not an average.')
    L.append('- The seed-collaboration scoreboard uses near-tie credit (0.3% '
             'relative, nearest-match): shares are distributional evidence and '
             'can sum to more than 100%.')
    L.append('- Never compare HV_fixed values across differently-configured '
             'campaigns: the frame is this run\'s own pooled ideal/nadir.')
    if exp.is_powercap:
        L.append('- Per-tier sections use the TIER frame: reference front and '
                 'bounds come from that tier\'s arms only. Tier-frame HV_fixed is '
                 'NOT comparable across tiers.')
        L.append('- For cross-cap comparisons of one arm, use the run-wide-frame '
                 'HV_fixed table (appendix of bundle.md): one common frame pooled '
                 'over all tiers.')
        L.append('- The global universal front mixes capped and uncapped regimes; '
                 'treat it as an envelope, never as a per-cap reference set.')
        L.append('- feasibility_summary.csv rates each arm\'s solutions against '
                 'every cap; uncapped arms land near the 90/60/30% calibration '
                 'targets by construction.')
    L.append('')
    L.append('## What to do')
    L.append('')
    steps = [
        'Rank the arms per scenario' + (' and per cap tier' if exp.is_powercap else '')
        + ': mean ± std HV_fixed first, then GD/IGD for convergence/coverage; '
          'check ranking stability across seeds (the per-seed tables) before '
          'claiming any ordering.',
        'Assess collaboration: which arms supply the universal front? Compare '
        'the strict pooled contribution (winner-takes-all) against the per-seed '
        'near-tie shares (distributional); a large divergence between the two '
        'is itself a finding worth reporting.',
    ]
    if exp.is_powercap:
        steps += [
            'Quantify each arm\'s degradation across caps: how much waiting time '
            'does each tier cost, what happens to energy, and how does run-wide '
            'HV_fixed fall from Uncapped to the tightest cap? Use the compare '
            'figures and the run-wide HV_fixed table.',
            'Sanity-check the calibration: uncapped arms\' feasibility under each '
            'cap should sit near its target percent; flag arms that deviate '
            'strongly (they run systematically hotter or cooler than the pool).',
        ]
    steps += [
        'Flag anomalies: arms with zero contribution everywhere, seeds that are '
        'outliers for one arm only, empty or one-point fronts, HV_fixed ties '
        'that flip under GD/IGD.',
        'End with a findings memo: ranked results with effect sizes (not just '
        'orderings), the collaboration picture, anomalies, and an explicit '
        'DO-NOT-CONCLUDE list (no cross-campaign HV comparisons; no legacy-HV '
        'citations; near-tie shares are shared credit'
        + ('; no cross-tier tier-frame HV comparisons' if exp.is_powercap else '')
        + ').',
    ]
    for i, s in enumerate(steps, 1):
        L.append(f'{i}. {s}')
    L.append('')
    L.append('The figures mirror the tables — use them to verify claims '
             'visually before asserting them numerically.')
    L.append('')
    return '\n'.join(L)


# =============================================================================
# EXCEL EXPORT — optional openpyxl dependency; everything else works without it
# =============================================================================

OPENPYXL_HINT = ('Excel export needs the openpyxl package:\n\n'
                 '    pip install openpyxl\n\n'
                 'CSV export works without it.')
XLSX_AGG_FILL = 'DDE8F0'   # the blue-grey the GUI uses for MEAN/STDDEV rows
XLSX_HEAD_FILL = 'F3F2F0'


def have_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def _xlsx_number_format(col):
    if col in ('Front size', 'Contribution', 'Runtime (ms)', 'TimeMs',
               'NonDomSolutions', 'ParetoContribution'):
        return '0'
    if col == '% of universal':
        return '0.0'
    if col == 'Avg runtime (s)':
        return '0.00'
    return '0.000000'


def _write_sheet(ws, df, info, display_map=None):
    """One table DataFrame into one worksheet, formatted like the GUI: bold
    frozen header row, MEAN/STDDEV/Universal rows shaded blue-grey, numeric
    cells with column-appropriate formats, auto-sized columns."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    display_map = display_map or {}
    head_font = Font(bold=True)
    head_fill = PatternFill('solid', fgColor=XLSX_HEAD_FILL)
    agg_fill = PatternFill('solid', fgColor=XLSX_AGG_FILL)
    agg_font = Font(bold=True)

    first_col = df.index.name or 'Seed'
    header = [first_col] + [display_map.get(c, c) for c in df.columns]
    ws.append(header)
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
    widths = [len(str(h)) for h in header]

    for idx, row in df.iterrows():
        agg = idx in ('MEAN', 'STDDEV', UNIVERSAL_KEY)
        name = (UNIVERSAL_LABEL if idx == UNIVERSAL_KEY
                else display_map.get(idx, idx))
        ws.append([name] + [None if pd.isna(v) else float(v) for v in row])
        r = ws.max_row
        widths[0] = max(widths[0], len(str(name)))
        for ci, col in enumerate(df.columns, start=2):
            ws.cell(row=r, column=ci).number_format = _xlsx_number_format(col)
        if agg:
            for ci in range(1, len(header) + 1):
                cell = ws.cell(row=r, column=ci)
                cell.fill = agg_fill
                cell.font = agg_font
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2, 11), 34)
    ws.freeze_panes = 'B2'
    if info:
        info_font = Font(italic=True, color='808080')
        ws.append([])
        for line in info.split('\n'):
            ws.append([line])
            ws.cell(row=ws.max_row, column=1).font = info_font


# Sheet-name suffix per table kind (Excel caps sheet names at 31 chars).
XLSX_SHEET_SUFFIX = {
    'Hypervolume': 'Hypervolume', 'GD': 'GD', 'IGD': 'IGD',
    'Spacing': 'Spacing', 'Front size': 'FrontSize',
    'Pareto contribution': 'Contribution', 'Runtime (ms)': 'Runtime',
}


def iter_export_tables(exp):
    """Yields (sheet_name, df, info) for every exportable table: per scenario
    — and per cap tier in PowerCap folders — the Global metrics table plus
    each per-seed pivot."""
    powercap = exp.is_powercap
    algos = exp.base_algorithms if powercap else exp.algorithms
    for n in sorted(exp.scenarios):
        scn = exp.scenarios[n]
        sources = ([(t, scn.tiers[t]) for t in exp.tier_names if t in scn.tiers]
                   if powercap else [(None, scn)])
        for tier, src in sources:
            prefix = f'S{n}' + (f' {tier}' if tier else '')
            df, info = global_metric_table(src, scn.obj_names, algos)
            if df is not None:
                yield f'{prefix} Global', df, info
            if src.metrics_seed is None:
                continue
            for disp, col in PER_SEED_TABLES:
                if col not in src.metrics_seed.columns:
                    continue
                piv, pinfo = pivot_metric_table(src, col)
                if piv is not None:
                    yield f'{prefix} {XLSX_SHEET_SUFFIX[disp]}', piv, pinfo


def _readme_lines(exp):
    objs = exp.objective_names
    first = next(iter(exp.scenarios.values()))
    n_seeds = (first.metrics_seed['Seed'].nunique()
               if first.metrics_seed is not None else 0)
    lines = [
        f'Experiment: {exp.name}',
        f'Study: {objs[0]} vs {objs[1]} — both objectives are minimized.',
        'Scenarios: ' + '; '.join(f'{n} = {exp.scenarios[n].name}'
                                  for n in sorted(exp.scenarios)),
        f'Seeds per (algorithm, scenario): {n_seeds}.',
        '',
        'How to read the sheets',
        '',
        'Global sheets: one row per algorithm — indicators of its COLLECTIVE '
        'Pareto set, the union of all its seed runs (non-dominated).',
        'Per-seed sheets: one row per seed, one column per algorithm, plus '
        'MEAN/STDDEV rows.',
        '',
        'Hypervolume: fixed-reference hypervolume. Fronts are normalized in '
        'the pooled ideal/nadir frame of the scenario and measured against '
        'the reference point at nadir + 10% of range on both objectives '
        '(= 1.1, 1.1 normalized). Higher is better, range 0..1. Comparable '
        'across algorithms and seeds within this run only — never across '
        'differently-configured campaigns.',
        'GD / IGD: average normalized distance from the front to the '
        'universal front / from the universal front to the front. '
        'Lower is better (convergence / coverage).',
        'Spacing: how evenly the front\'s points are spread (lower = more '
        'even); says nothing about quality.',
        'Front size: number of points on the (non-dominated) front.',
        'Pareto contribution: how many universal-front points the algorithm '
        'matched exactly (1e-9 tolerance). On per-seed sheets the MEAN row '
        'is the union over seeds, not an average.',
        'Runtime: per-seed sheets are milliseconds; Global sheets show the '
        'average over the seed runs, in seconds.',
    ]
    if exp.is_powercap:
        lines += ['',
                  'PowerCeiling study: sheets are per cap tier. Every number '
                  'is computed within that tier\'s own frame and against the '
                  'tier\'s own universal front, so Hypervolume must not be '
                  'compared across tiers.']
    if exp.hv_fixed_recomputed:
        lines += ['',
                  'Note: this folder pre-dates the HV_fixed CSV column; the '
                  'Hypervolume values were computed by the explorer with the '
                  'same fixed-reference formula.']
    return lines


def export_tables_xlsx(exp, path):
    """Every table into one workbook: a README sheet plus one sheet per table
    (see iter_export_tables). Requires openpyxl — callers catch ImportError
    and show OPENPYXL_HINT. Returns the number of table sheets written."""
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'README'
    for line in _readme_lines(exp):
        ws.append([line] if line else [])
    ws.column_dimensions['A'].width = 100
    for row in (1, 6):
        ws.cell(row=row, column=1).font = Font(bold=True)
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).alignment = \
            openpyxl.styles.Alignment(wrap_text=True, vertical='top')

    display_map = {a: st['display'] for a, st in resolve_styles(
        exp.base_algorithms if exp.is_powercap else exp.algorithms).items()}
    count = 0
    for name, df, info in iter_export_tables(exp):
        _write_sheet(wb.create_sheet(title=name[:31]), df, info, display_map)
        count += 1
    wb.save(path)
    return count


def write_single_table_xlsx(df, info, path, display_map=None, sheet='Table'):
    """The currently shown table as a one-sheet workbook (same formatting)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = re.sub(r'[\[\]:*?/\\]', ' ', sheet)[:31] or 'Table'
    _write_sheet(ws, df, info, display_map)
    wb.save(path)


# =============================================================================
# HEADLESS BULK EXPORT (also exercises every builder without a display)
# =============================================================================

# Per-seed pivots emitted by the bulk exports (single HV; no legacy columns).
TABLE_METRICS = ('HV_fixed', 'GD', 'IGD', 'Spacing', 'NonDomSolutions',
                 'ParetoContribution', 'TimeMs')


def _write_table(path, piv, info):
    with open(path, 'w') as fh:
        piv.to_csv(fh, index_label=piv.index.name or 'Seed',
                   float_format='%.6f')
        if info:
            fh.write(f'\n# {info}\n')
    return path


def _emit_figures(exp, outdir, dpi, described):
    """Renders every figure the GUI can draw into outdir; appends
    (path, description) pairs to described. Powercap folders get per-tier
    scatters + bars, per-arm compare figures and the feasibility chart."""
    styles = resolve_styles(exp.base_algorithms if exp.is_powercap
                            else exp.algorithms)
    base_opts = {'show_legend': True}

    def save(fig, fname, desc):
        p = os.path.join(outdir, fname)
        fig.savefig(p, dpi=dpi)
        described.append((p, desc))

    for n in sorted(exp.scenarios):
        scn = exp.scenarios[n]
        if exp.is_powercap:
            for tier in exp.tier_names:
                opts = dict(base_opts, mode='cap', tier=tier)
                fig = build_scatter_figure(exp, scn, styles, opts)
                save(fig, f'scatter_scenario_{n}_{tier.lower()}.png',
                     f'scenario {n} fronts at {tier_display(tier)} '
                     '(+ tier universal front)')
            for base in exp.base_algorithms:
                opts = dict(base_opts, mode='cmp', cmp_algo=base)
                fig = build_compare_figure(exp, scn, base, styles, opts)
                save(fig, f'compare_{base}_scenario_{n}.png',
                     f'{base} across all cap tiers, scenario {n}')
        else:
            fig = build_scatter_figure(exp, scn, styles, dict(base_opts))
            save(fig, f'scatter_scenario_{n}.png',
                 f'scenario {n} fronts + universal front')

    aggs = (('collective', 'collective pooled fronts'),
            ('mean', 'mean ± std over seeds'))
    if exp.is_powercap:
        for tier in exp.tier_names:
            for key, display, ylabel in powercap_metrics(exp):
                for agg, agg_desc in aggs:
                    opts = dict(base_opts, mode='cap', tier=tier,
                                aggregation=agg)
                    fig = build_metric_figure(exp, key, display, ylabel, None,
                                              styles, opts)
                    save(fig, f'bars_{key}_{tier.lower()}_{agg}.png',
                         f'{display} ({agg_desc}), all scenarios, '
                         f'{tier_display(tier)}')
            fig = build_runtime_figure(exp, styles,
                                       dict(base_opts, mode='cap', tier=tier))
            save(fig, f'bars_runtime_{tier.lower()}.png',
                 f'runtime per arm at {tier_display(tier)}')
        fig = build_feasibility_figure(exp, styles, dict(base_opts))
        save(fig, 'feasibility.png',
             'uncapped arms\' feasibility under each derived cap vs the '
             'calibration targets')
    else:
        for key, display, ylabel in available_metrics(exp):
            for agg, agg_desc in aggs:
                fig = build_metric_figure(exp, key, display, ylabel, None,
                                          styles,
                                          dict(base_opts, aggregation=agg))
                save(fig, f'bars_{key}_{agg}.png',
                     f'{display} ({agg_desc}), all scenarios')
        fig = build_runtime_figure(exp, styles, dict(base_opts))
        save(fig, 'bars_runtime.png', 'runtime per algorithm')


def _emit_tables(exp, outdir, described, include_collab=True):
    """Writes every table as a real CSV into outdir (per tier for powercap
    folders): the Global metrics table + every per-seed pivot, plus
    powercap-only copies of the universal fronts and the feasibility summary.
    include_collab keeps the near-tie collaboration CSVs (Claude zip only —
    the GUI-facing exports leave them out by owner decision)."""

    def emit_source(src, algos, obj_names, stem, where):
        gdf, ginfo = global_metric_table(src, obj_names, algos)
        if gdf is not None:
            p = _write_table(os.path.join(outdir, f'{stem}_global_metrics.csv'),
                             gdf, ginfo)
            described.append((p, f'Global metrics (collective pooled fronts), '
                                 f'{where}'))
        for metric in TABLE_METRICS:
            piv, info = pivot_metric_table(src, metric)
            if piv is None:
                continue
            p = _write_table(os.path.join(outdir, f'{stem}_{metric}.csv'),
                             piv, info)
            described.append((p, f'seeds × algorithms pivot of {metric}, '
                                 f'{where}'))
        if include_collab:
            piv, info = collab_table(src)
            if piv is not None:
                p = _write_table(
                    os.path.join(outdir, f'{stem}_seed_collaboration.csv'),
                    piv, info)
                described.append((p, f'near-tie collaboration shares, {where}'))

    for n in sorted(exp.scenarios):
        scn = exp.scenarios[n]
        if exp.is_powercap:
            for tier in exp.tier_names:
                td = scn.tiers.get(tier)
                if td is None:
                    continue
                emit_source(td, exp.base_algorithms, scn.obj_names,
                            f'table_scenario_{n}_{tier.lower()}',
                            f'scenario {n}, {tier_display(tier)}')
            # Universal fronts of every tier (works for native and recomputed).
            rows = []
            for tier in exp.tier_names:
                td = scn.tiers.get(tier)
                if td is None or td.universal is None:
                    continue
                block = td.universal[scn.obj_names].copy()
                block.insert(0, 'CapTier', tier)
                rows.append(block)
            if rows:
                p = os.path.join(outdir, f'scenario_{n}_universal_fronts_by_cap.csv')
                pd.concat(rows, ignore_index=True).to_csv(p, index=False,
                                                          float_format='%.9f')
                described.append((p, f'every tier\'s universal front, scenario {n}'))
        else:
            emit_source(scn, exp.algorithms, scn.obj_names,
                        f'table_scenario_{n}', f'scenario {n}')

    if exp.is_powercap and exp.feasibility is not None:
        p = os.path.join(outdir, 'feasibility_summary.csv')
        exp.feasibility.to_csv(p, index=False)
        described.append((p, 'per (scenario, arm, cap) feasibility rates '
                             '(copy of the campaign file)'))


def export_all(exp, outdir, dpi=300):
    """Headless export of every figure/table + the markdown bundle into a
    directory (the --export-all path). Same emitters feed the Claude zip."""
    os.makedirs(outdir, exist_ok=True)
    described = []
    _emit_figures(exp, outdir, dpi, described)
    _emit_tables(exp, outdir, described, include_collab=False)
    p = os.path.join(outdir, f'claude_bundle_{exp.name}.md')
    with open(p, 'w') as fh:
        fh.write(build_claude_bundle(exp))
    described.append((p, 'all tables inline as markdown'))
    return [p for p, _ in described]


def export_claude_zip(exp, zip_path, dpi=200):
    """One self-briefing zip for LLM analysis: INSTRUCTIONS_FOR_CLAUDE.md
    (generated), bundle.md (all tables inline), figures/ (every plot) and
    tables/ (every pivot as a real CSV). Works for every study; powercap
    folders add the tier dimension and powercap-only files."""
    with tempfile.TemporaryDirectory() as tmp:
        fig_dir = os.path.join(tmp, 'figures')
        tab_dir = os.path.join(tmp, 'tables')
        os.makedirs(fig_dir)
        os.makedirs(tab_dir)
        described = []
        _emit_figures(exp, fig_dir, dpi, described)
        # The zip keeps the near-tie collaboration CSVs: it is LLM input and
        # its briefing explains them (the GUI/Excel exports leave them out).
        _emit_tables(exp, tab_dir, described, include_collab=True)

        bundle_path = os.path.join(tmp, 'bundle.md')
        with open(bundle_path, 'w') as fh:
            fh.write(build_claude_bundle(exp))

        file_map = [('INSTRUCTIONS_FOR_CLAUDE.md', 'this file — read it first'),
                    ('bundle.md', 'every table inline as markdown CSV blocks '
                                  '(self-contained fallback if you can only '
                                  'read one file)')]
        file_map += [(os.path.relpath(p, tmp), desc) for p, desc in described]
        with open(os.path.join(tmp, 'INSTRUCTIONS_FOR_CLAUDE.md'), 'w') as fh:
            fh.write(build_instructions_md(exp, file_map))

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, _dirs, files in os.walk(tmp):
                for fname in sorted(files):
                    full = os.path.join(root, fname)
                    zf.write(full, os.path.relpath(full, tmp))
    return zip_path

# =============================================================================
# GUI — Tkinter is imported only inside launch_gui() so headless use
# (--export-all) works on machines without a display or python3-tk.
# =============================================================================

PLOT_TABS = ('scatter', 'metrics', 'runtime', 'feas')


class ResultsExplorerApp:
    """Main window. Composition over tk.Tk so the module imports headlessly."""

    def __init__(self, tkmod, folder=None):
        # tkmod bundles the lazily imported tkinter modules (see launch_gui).
        self.tk = tkmod['tk']
        self.ttk = tkmod['ttk']
        self.filedialog = tkmod['filedialog']
        self.colorchooser = tkmod['colorchooser']
        self.messagebox = tkmod['messagebox']
        self.FigureCanvasTkAgg = tkmod['FigureCanvasTkAgg']
        self.NavigationToolbar2Tk = tkmod['NavigationToolbar2Tk']

        self.exp = None
        self.styles = {}
        self.hidden = set()
        self.custom_titles = {}       # tab key -> custom title ('' = default)
        self.figures = {}             # tab key -> current Figure
        self.canvases = {}            # tab key -> FigureCanvasTkAgg
        self.current_table = None     # (DataFrame, info) shown in Tables tab
        self._details_sols = []       # solutions of the current details filter

        root = self.tk.Tk()
        self.root = root
        root.title('COSMOS Results Explorer')
        root.geometry('1420x880')
        self._build_ui()
        if folder:
            self.load_folder(folder)

    def run(self):
        self.root.mainloop()

    # ---- UI skeleton -----------------------------------------------------

    def _build_ui(self):
        tk, ttk = self.tk, self.ttk

        top = ttk.Frame(self.root, padding=(8, 6))
        top.pack(side='top', fill='x')
        ttk.Button(top, text='Open experiment…', command=self.open_dialog).pack(side='left')
        self.exp_label = ttk.Label(top, text='No experiment loaded', foreground='#666')
        self.exp_label.pack(side='left', padx=12)

        body = ttk.Frame(self.root)
        body.pack(fill='both', expand=True)

        side = ttk.Frame(body, padding=(8, 4), width=260)
        side.pack(side='left', fill='y')
        side.pack_propagate(False)
        self._build_sidebar(side)

        main = ttk.Frame(body)
        main.pack(side='left', fill='both', expand=True)
        self.notebook = ttk.Notebook(main)
        self.notebook.pack(fill='both', expand=True, padx=4, pady=4)
        self.tabs = {}
        for key, title in (('scatter', 'Scatter'), ('metrics', 'Metrics'),
                           ('runtime', 'Runtime'), ('tables', 'Tables'),
                           ('details', 'Details'), ('feas', 'Feasibility')):
            frame = ttk.Frame(self.notebook)
            self.notebook.add(frame, text=title)
            self.tabs[key] = frame
        self.notebook.hide(self.tabs['feas'])  # powercap folders only
        self.notebook.bind('<<NotebookTabChanged>>', lambda e: self._on_tab_changed())

        self._build_metrics_controls()
        self._build_tables_tab()
        self._build_details_tab()

        self.status = ttk.Label(self.root, text='Open a results/<ExperimentId> folder to begin.',
                                anchor='w', padding=(8, 3))
        self.status.pack(side='bottom', fill='x')

    def _sideheader(self, parent, text):
        self.ttk.Label(parent, text=text.upper(), foreground='#888',
                       font=('TkDefaultFont', 8, 'bold')).pack(anchor='w', pady=(10, 2))

    def _build_sidebar(self, side):
        tk, ttk = self.tk, self.ttk

        def header(text):
            self._sideheader(side, text)

        header('Scenario')
        self.scenario_var = tk.IntVar(value=1)
        self.scenario_frame = ttk.Frame(side)
        self.scenario_frame.pack(anchor='w', fill='x')

        # Everything between Scenario and Display is study-shaped: for powercap
        # folders it carries the view switch, cap dropdown / tier list, and the
        # Pareto-set toggles. Rebuilt by _rebuild_dynamic_sidebar.
        self.dyn_frame = ttk.Frame(side)
        self.dyn_frame.pack(anchor='w', fill='x')

        # PowerCap state (harmless defaults for ordinary folders).
        self.mode_var = tk.StringVar(value='cap')
        self.tier_var = tk.StringVar(value='')
        self.cmp_var = tk.StringVar(value='')
        self._tier_vars = {}
        self.show_universal_var = tk.BooleanVar(value=True)
        self.show_global_universal_var = tk.BooleanVar(value=False)

        header('Display')
        self.legend_var = tk.BooleanVar(value=True)
        self.labels_var = tk.BooleanVar(value=False)
        self.clouds_var = tk.BooleanVar(value=False)
        self.swap_var = tk.BooleanVar(value=False)
        self.normalize_var = tk.BooleanVar(value=False)
        for var, text in ((self.legend_var, 'Show legend'),
                          (self.labels_var, 'Algorithm name labels'),
                          (self.clouds_var, 'Per-seed point clouds'),
                          (self.swap_var, 'Swap X / Y axes'),
                          (self.normalize_var, 'Normalize axes (pooled ideal/nadir)')):
            ttk.Checkbutton(side, text=text, variable=var,
                            command=self.refresh_plots).pack(anchor='w')
        self.ref_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(side, text='Show HV reference point', variable=self.ref_var,
                        command=self._ref_changed).pack(anchor='w')

        ttk.Label(side, text='Title (empty = default):').pack(anchor='w', pady=(8, 0))
        self.title_entry = ttk.Entry(side)
        self.title_entry.pack(fill='x')
        self.title_entry.bind('<Return>', lambda e: self._apply_title())
        self.title_entry.bind('<FocusOut>', lambda e: self._apply_title())

        row = ttk.Frame(side)
        row.pack(anchor='w', pady=(8, 0))
        ttk.Label(row, text='Marker').pack(side='left')
        self.marker_var = tk.StringVar(value='7')
        tk.Spinbox(row, from_=2, to=20, width=4, textvariable=self.marker_var,
                   command=self.refresh_plots).pack(side='left', padx=(4, 10))
        ttk.Label(row, text='Save DPI').pack(side='left')
        self.dpi_var = tk.StringVar(value='300')
        tk.Spinbox(row, from_=72, to=600, increment=25, width=5,
                   textvariable=self.dpi_var).pack(side='left', padx=4)

        actions = ttk.Frame(side)
        actions.pack(side='bottom', fill='x', pady=8)
        for text, cmd in (('Save figure…', self.save_figure),
                          ('Save table…', self.save_table),
                          ('Export all…', self.export_all_dialog),
                          ('Save for Claude', self.save_for_claude)):
            ttk.Button(actions, text=text, command=cmd).pack(fill='x', pady=2)

    def _build_metrics_controls(self):
        ttk = self.ttk
        bar = ttk.Frame(self.tabs['metrics'])
        bar.pack(side='top', fill='x', padx=4, pady=4)
        ttk.Label(bar, text='Indicator:').pack(side='left')
        self.metric_var = self.tk.StringVar()
        self.metric_combo = ttk.Combobox(bar, textvariable=self.metric_var,
                                         state='readonly', width=44)
        self.metric_combo.pack(side='left', padx=6)
        self.metric_combo.bind('<<ComboboxSelected>>',
                               lambda e: self._refresh_tab('metrics'))
        ttk.Label(bar, text='Aggregation:').pack(side='left', padx=(14, 0))
        self.agg_var = self.tk.StringVar(value='collective')
        self._agg_buttons = []
        for value, text in (('collective', 'Collective (pooled 10-seed front)'),
                            ('mean', 'Mean ± std over seeds')):
            rb = ttk.Radiobutton(bar, text=text, value=value,
                                 variable=self.agg_var,
                                 command=lambda: self._refresh_tab('metrics'))
            rb.pack(side='left', padx=(6, 0))
            self._agg_buttons.append(rb)

    def _build_tables_tab(self):
        ttk = self.ttk
        frame = self.tabs['tables']
        bar = ttk.Frame(frame)
        bar.pack(side='top', fill='x', padx=4, pady=4)
        ttk.Label(bar, text='Table:').pack(side='left')
        self.table_kind_var = self.tk.StringVar()
        self.table_kind_combo = ttk.Combobox(bar, textvariable=self.table_kind_var,
                                             state='readonly', width=44)
        self.table_kind_combo.pack(side='left', padx=6)
        self.table_kind_combo.bind('<<ComboboxSelected>>',
                                   lambda e: self._refresh_tab('tables'))
        ttk.Button(bar, text='Export all tables → Excel…',
                   command=self.export_workbook_dialog).pack(side='right', padx=2)
        ttk.Button(bar, text='Export table → Excel',
                   command=self.export_table_xlsx).pack(side='right', padx=2)

        holder = ttk.Frame(frame)
        holder.pack(fill='both', expand=True, padx=4)
        self.table_tree = ttk.Treeview(holder, show='headings')
        ysb = ttk.Scrollbar(holder, orient='vertical', command=self.table_tree.yview)
        xsb = ttk.Scrollbar(holder, orient='horizontal', command=self.table_tree.xview)
        self.table_tree.configure(yscroll=ysb.set, xscroll=xsb.set)
        self.table_tree.grid(row=0, column=0, sticky='nsew')
        ysb.grid(row=0, column=1, sticky='ns')
        xsb.grid(row=1, column=0, sticky='ew')
        holder.rowconfigure(0, weight=1)
        holder.columnconfigure(0, weight=1)
        self.table_tree.tag_configure('agg', background='#dde8f0')

        self.table_info = ttk.Label(frame, text='', foreground='#666', padding=(6, 4))
        self.table_info.pack(anchor='w')

    def _build_details_tab(self):
        ttk = self.ttk
        frame = self.tabs['details']

        self.details_msg = ttk.Label(
            frame, padding=14, foreground='#666', wraplength=900, text=(
                'No solution details in this run.\n\n'
                'scenario_N_solution_details.json.gz is written by campaigns run '
                'on a build that includes the SolutionDetailsCollector exporter '
                '(every published Pareto-front solution is captured). Re-run the '
                'campaign to get this view.'))

        self.details_body = ttk.Frame(frame)

        bar = ttk.Frame(self.details_body)
        bar.pack(side='top', fill='x', padx=4, pady=4)
        self.d_algo_var = self.tk.StringVar()
        self.d_seed_var = self.tk.StringVar()
        self.d_sol_var = self.tk.StringVar()
        ttk.Label(bar, text='Algorithm:').pack(side='left')
        self.d_algo_combo = ttk.Combobox(bar, textvariable=self.d_algo_var,
                                         state='readonly', width=26)
        self.d_algo_combo.pack(side='left', padx=(2, 8))
        ttk.Label(bar, text='Seed:').pack(side='left')
        self.d_seed_combo = ttk.Combobox(bar, textvariable=self.d_seed_var,
                                         state='readonly', width=7)
        self.d_seed_combo.pack(side='left', padx=(2, 8))
        ttk.Label(bar, text='Solution:').pack(side='left')
        self.d_sol_combo = ttk.Combobox(bar, textvariable=self.d_sol_var,
                                        state='readonly', width=52)
        self.d_sol_combo.pack(side='left', padx=2, fill='x', expand=True)
        self.d_algo_combo.bind('<<ComboboxSelected>>', lambda e: self._details_filter_changed())
        self.d_seed_combo.bind('<<ComboboxSelected>>', lambda e: self._details_filter_changed())
        self.d_sol_combo.bind('<<ComboboxSelected>>', lambda e: self._details_show_selected())

        grids = ttk.Frame(self.details_body)
        grids.pack(fill='both', expand=True, padx=4, pady=4)
        grids.columnconfigure(0, weight=1)
        grids.columnconfigure(1, weight=1)
        grids.rowconfigure(1, weight=1)
        grids.rowconfigure(3, weight=2)

        def make_tree(parent, columns, widths):
            holder = ttk.Frame(parent)
            tree = ttk.Treeview(holder, show='headings', columns=columns)
            for col, w in zip(columns, widths):
                tree.heading(col, text=col)
                tree.column(col, width=w, anchor='e' if col not in
                            ('Queue (dispatch order)', 'Type', 'Workload') else 'w')
            ysb = ttk.Scrollbar(holder, orient='vertical', command=tree.yview)
            tree.configure(yscroll=ysb.set)
            tree.grid(row=0, column=0, sticky='nsew')
            ysb.grid(row=0, column=1, sticky='ns')
            holder.rowconfigure(0, weight=1)
            holder.columnconfigure(0, weight=1)
            return holder, tree

        ttk.Label(grids, text='Solution summary').grid(row=0, column=0, sticky='w')
        holder, self.d_summary_tree = make_tree(
            grids, ('Metric', 'Value'), (220, 160))
        holder.grid(row=1, column=0, sticky='nsew', padx=(0, 6))
        self.d_summary_tree.column('Metric', anchor='w')

        ttk.Label(grids, text='Per-host energy').grid(row=0, column=1, sticky='w')
        holder, self.d_host_tree = make_tree(
            grids, ('Host', 'Type', 'Cores', 'GPUs', 'Energy (kWh)'),
            (50, 130, 55, 50, 120))
        holder.grid(row=1, column=1, sticky='nsew')
        self.d_host_tree.column('Type', anchor='w')

        ttk.Label(grids, text='Per-VM queues').grid(row=2, column=0, sticky='w', pady=(8, 0))
        holder, self.d_vm_tree = make_tree(
            grids, ('VM', 'Type', 'vCPUs', 'GPUs', 'Host', 'Queue size',
                    'Queue (dispatch order)'),
            (44, 70, 55, 48, 48, 80, 420))
        holder.grid(row=3, column=0, sticky='nsew', padx=(0, 6))

        ttk.Label(grids, text='Task assignments').grid(row=2, column=1, sticky='w', pady=(8, 0))
        holder, self.d_task_tree = make_tree(
            grids, ('Task', 'Length (instr.)', 'Workload', '→ VM', 'Wait (s)'),
            (55, 110, 130, 55, 80))
        holder.grid(row=3, column=1, sticky='nsew')

        self.details_msg.pack(fill='both', expand=True)

    # ---- loading -----------------------------------------------------------

    def open_dialog(self):
        folder = self.filedialog.askdirectory(
            title='Select an experiment results folder',
            initialdir=self._initial_dir())
        if folder:
            self.load_folder(folder)

    def _initial_dir(self):
        here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        for cand in (os.path.join(here, 'results'),
                     os.path.join(here, 'newExperimentResults'), here):
            if os.path.isdir(cand):
                return cand
        return os.getcwd()

    def load_folder(self, folder):
        try:
            exp = ExperimentData.load(folder)
        except Exception as e:
            self.messagebox.showerror('Could not load experiment', str(e))
            return
        self.exp = exp
        self.styles = resolve_styles(
            exp.base_algorithms if exp.is_powercap else exp.algorithms)
        self.hidden = set()
        self.custom_titles = {}
        self.mode_var.set('cap')
        self.tier_var.set(exp.tier_names[0] if exp.is_powercap else '')
        self.cmp_var.set(exp.base_algorithms[0] if exp.is_powercap else '')
        self._tier_vars = {}
        self.show_universal_var.set(True)
        self.show_global_universal_var.set(False)
        objs = exp.objective_names
        n_scen = len(exp.scenarios)
        n_seeds = 0
        first = next(iter(exp.scenarios.values()))
        if first.metrics_seed is not None:
            n_seeds = first.metrics_seed['Seed'].nunique()
        if exp.is_powercap:
            arms = (f'{len(exp.base_algorithms)} base arms × '
                    f'{len(exp.tier_names)} cap tiers')
        else:
            arms = f'{len(exp.algorithms)} algorithms'
        self.exp_label.config(text=(
            f'{exp.name}   ·   {objs[0]} vs {objs[1]} · {n_scen} scenarios · '
            f'{arms} · {n_seeds} seeds'))
        if exp.is_powercap:
            self.notebook.add(self.tabs['feas'], text='Feasibility')
        else:
            self.notebook.hide(self.tabs['feas'])
        self._rebuild_scenario_controls()
        self._rebuild_dynamic_sidebar()
        self._rebuild_metric_choices()
        self._rebuild_table_choices()
        self.refresh_all()
        note = (' · Hypervolume computed in-explorer (folder pre-dates the '
                'HV_fixed column)' if exp.hv_fixed_recomputed else '')
        self.set_status(f'Loaded {exp.folder}{note}')
        self._update_powercap_status()

    def _rebuild_scenario_controls(self):
        for w in self.scenario_frame.winfo_children():
            w.destroy()
        nums = sorted(self.exp.scenarios)
        self.scenario_var.set(nums[0])
        for n in nums:
            self.ttk.Radiobutton(
                self.scenario_frame, text=f'{n} — {self.exp.scenarios[n].name}',
                value=n, variable=self.scenario_var,
                command=self._scenario_changed).pack(anchor='w')

    def _scenario_changed(self):
        self._refresh_tab('scatter')
        self._refresh_tab('tables')
        self._refresh_details_tab()

    def _rebuild_dynamic_sidebar(self):
        """(Re)builds the study-shaped sidebar block: view switch + cap dropdown
        + algorithm list + Pareto-set toggles for powercap folders (per mode),
        or the plain algorithm list + universal toggle otherwise. Preserves the
        current visibility/color state where labels persist."""
        tk, ttk = self.tk, self.ttk
        dyn = self.dyn_frame
        for w in dyn.winfo_children():
            w.destroy()
        exp = self.exp
        if exp is None:
            return
        powercap = exp.is_powercap
        mode = self.mode_var.get() if powercap else 'cap'

        if powercap:
            self._sideheader(dyn, 'View')
            for value, text in (('cap', 'Per-cap view'),
                                ('cmp', 'Compare algorithm across caps')):
                ttk.Radiobutton(dyn, text=text, value=value, variable=self.mode_var,
                                command=self._mode_changed).pack(anchor='w')

        if powercap and mode == 'cap':
            self._sideheader(dyn, 'Power cap')
            self.tier_combo = ttk.Combobox(dyn, state='readonly', width=26)
            tiers = exp.tier_names
            self.tier_combo['values'] = [
                tier_display(t, exp.cap_watts.get(t)) for t in tiers]
            if self.tier_var.get() not in tiers:
                self.tier_var.set(tiers[0])
            self.tier_combo.current(tiers.index(self.tier_var.get()))
            self.tier_combo.pack(anchor='w', fill='x')
            self.tier_combo.bind('<<ComboboxSelected>>', lambda e: self._tier_changed())

        if mode == 'cap':
            self._sideheader(dyn, 'Algorithms — color & visibility')
            self.algo_frame = ttk.Frame(dyn)
            self.algo_frame.pack(anchor='w', fill='x')
            self._algo_vars = {}
            algos = exp.base_algorithms if powercap else exp.algorithms
            for algo in algos:
                st = self.styles[algo]
                row = ttk.Frame(self.algo_frame)
                row.pack(anchor='w', fill='x')
                var = tk.BooleanVar(value=algo not in self.hidden)
                self._algo_vars[algo] = var
                ttk.Checkbutton(row, variable=var,
                                command=self._visibility_changed).pack(side='left')
                btn = tk.Button(row, width=2, bg=st['color'],
                                activebackground=st['color'], relief='raised',
                                command=lambda a=algo: self.pick_color(a))
                btn.pack(side='left', padx=(0, 6))
                st['_swatch'] = btn
                ttk.Label(row, text=st['display']).pack(side='left')

            self._sideheader(dyn, 'Pareto sets')
            row = ttk.Frame(dyn)
            row.pack(anchor='w', fill='x')
            ttk.Checkbutton(
                row, variable=self.show_universal_var, command=self.refresh_plots,
                text=(f'Universal — {tier_display(self.tier_var.get())}'
                      if powercap else UNIVERSAL_LABEL)).pack(side='left')
            if powercap:
                row = ttk.Frame(dyn)
                row.pack(anchor='w', fill='x')
                ttk.Checkbutton(row, variable=self.show_global_universal_var,
                                command=self.refresh_plots,
                                text='Global universal (all arms)').pack(side='left')
                ttk.Label(dyn, foreground='#888', font=('TkDefaultFont', 8),
                          text='global = pooled over uncapped + all caps').pack(anchor='w')

        elif mode == 'cmp':
            self._sideheader(dyn, 'Algorithm')
            self.cmp_combo = ttk.Combobox(dyn, state='readonly', width=26)
            self.cmp_combo['values'] = [self.styles[a]['display']
                                        for a in exp.base_algorithms]
            if self.cmp_var.get() not in exp.base_algorithms:
                self.cmp_var.set(exp.base_algorithms[0])
            self.cmp_combo.current(exp.base_algorithms.index(self.cmp_var.get()))
            self.cmp_combo.pack(anchor='w', fill='x')
            self.cmp_combo.bind('<<ComboboxSelected>>', lambda e: self._cmp_changed())

            self._sideheader(dyn, 'Cap tiers — visibility')
            old = {t: v.get() for t, v in self._tier_vars.items()}
            self._tier_vars = {}
            for tier in exp.tier_names:
                row = ttk.Frame(dyn)
                row.pack(anchor='w', fill='x')
                var = tk.BooleanVar(value=old.get(tier, True))
                self._tier_vars[tier] = var
                ttk.Checkbutton(row, variable=var,
                                command=self.refresh_plots).pack(side='left')
                sw = tk.Frame(row, width=16, height=6, bg=tier_color(exp, tier))
                sw.pack(side='left', padx=(0, 6))
                text = tier_display(tier, exp.cap_watts.get(tier))
                if tier == UNCAPPED_TIER:
                    text += ' — dashed'
                ttk.Label(row, text=text).pack(side='left')

    def _visibility_changed(self):
        self.hidden = {a for a, v in self._algo_vars.items() if not v.get()}
        self.refresh_plots()

    def _ref_changed(self):
        # The reference point shows on the plots AND in the table info lines.
        self.refresh_plots()
        self._refresh_tab('tables')

    def _mode_changed(self):
        self._rebuild_dynamic_sidebar()
        self._rebuild_metric_choices()
        self._rebuild_table_choices()
        self.refresh_all()
        self._update_powercap_status()

    def _tier_changed(self):
        tiers = self.exp.tier_names
        self.tier_var.set(tiers[self.tier_combo.current()])
        self._rebuild_dynamic_sidebar()  # refresh the universal-toggle label
        self._rebuild_table_choices()
        self.refresh_all()
        self._update_powercap_status()

    def _cmp_changed(self):
        self.cmp_var.set(self.exp.base_algorithms[self.cmp_combo.current()])
        self.refresh_all()
        self._update_powercap_status()

    def _update_powercap_status(self):
        exp = self.exp
        if exp is None or not exp.is_powercap:
            return
        src = ('native (*_by_cap.csv)' if exp.by_cap_source == 'native'
               else 'recomputed in explorer (pre-fix folder)')
        if self.mode_var.get() == 'cmp':
            disp = self.styles.get(self.cmp_var.get(), {}).get('display', self.cmp_var.get())
            self.set_status(f'Compare mode: {disp} across '
                            f'{sum(v.get() for v in self._tier_vars.values())} cap tiers · '
                            f'HV frame: run-wide (comparable across caps) · per-tier data: {src}')
        else:
            tier = self.tier_var.get()
            self.set_status(f'{tier_display(tier, exp.cap_watts.get(tier))} · '
                            f'per-tier data: {src}')

    def pick_color(self, algo):
        st = self.styles[algo]
        rgb, hexcolor = self.colorchooser.askcolor(
            color=st['color'], title=f'Color for {st["display"]}')
        if hexcolor:
            st['color'] = hexcolor
            st['_swatch'].config(bg=hexcolor, activebackground=hexcolor)
            self.refresh_plots()

    def _rebuild_metric_choices(self):
        exp = self.exp
        compare = exp.is_powercap and self.mode_var.get() == 'cmp'
        if compare:
            choices = COMPARE_METRIC_CHOICES  # 4-tuples: (..., source)
        elif exp.is_powercap:
            choices = powercap_metrics(exp)
        else:
            choices = available_metrics(exp)
        self._metric_defs = {}
        for choice in choices:
            key, display, ylabel = choice[:3]
            self._metric_defs[display] = (
                key, ylabel, choice[3] if len(choice) > 3 else None)
        values = list(self._metric_defs)
        self.metric_combo['values'] = values
        if values:
            self.metric_var.set(values[0])
        # Compare mode has its own cross-cap semantics — no aggregation choice.
        for rb in self._agg_buttons:
            rb.config(state='disabled' if compare else 'normal')

    def _rebuild_table_choices(self):
        exp = self.exp
        kinds = []
        if exp.is_powercap and self.mode_var.get() == 'cmp':
            kinds = list(COMPARE_TABLE_KINDS)
        else:
            src = next(iter(exp.scenarios.values()))
            if exp.is_powercap:
                src = src.tiers.get(self.tier_var.get(), src)
            kinds.append(GLOBAL_TABLE_KIND)
            if src.metrics_seed is not None:
                kinds += [f'Per-seed — {disp}' for disp, col in PER_SEED_TABLES
                          if col in src.metrics_seed.columns]
        self.table_kind_combo['values'] = kinds
        if kinds:
            self.table_kind_var.set(kinds[0])

    # ---- refresh -----------------------------------------------------------

    def set_status(self, text):
        self.status.config(text=text)

    def _current_tab(self):
        # Match by widget path (robust when tabs are hidden, e.g. Feasibility).
        sel = self.notebook.select()
        for key, frame in self.tabs.items():
            if str(frame) == sel:
                return key
        return 'scatter'

    def _on_tab_changed(self):
        tab = self._current_tab()
        self.title_entry.delete(0, 'end')
        self.title_entry.insert(0, self.custom_titles.get(tab, ''))

    def _apply_title(self):
        tab = self._current_tab()
        if tab in PLOT_TABS:
            self.custom_titles[tab] = self.title_entry.get().strip()
            self._refresh_tab(tab)

    def _opts(self, tab):
        try:
            ms = float(self.marker_var.get())
        except ValueError:
            ms = 7.0
        opts = {
            'swap': self.swap_var.get(),
            'title': self.custom_titles.get(tab, ''),
            'show_legend': self.legend_var.get(),
            'show_labels': self.labels_var.get(),
            'show_clouds': self.clouds_var.get(),
            'normalize': self.normalize_var.get(),
            'hidden': self.hidden,
            'marker_size': ms,
            'show_universal': self.show_universal_var.get(),
            'show_ref_point': self.ref_var.get(),
            'aggregation': self.agg_var.get(),
        }
        if self.exp is not None and self.exp.is_powercap:
            mode = self.mode_var.get()
            opts['mode'] = mode
            if mode == 'cap':
                opts['tier'] = self.tier_var.get() or self.exp.tier_names[0]
                opts['show_global_universal'] = self.show_global_universal_var.get()
            else:
                opts['cmp_algo'] = self.cmp_var.get()
                opts['tiers_hidden'] = {t for t, v in self._tier_vars.items()
                                        if not v.get()}
        return opts

    def refresh_all(self):
        self.refresh_plots()
        self._refresh_tab('tables')
        self._refresh_details_tab()

    def refresh_plots(self):
        for tab in PLOT_TABS:
            self._refresh_tab(tab)

    def _refresh_tab(self, tab):
        if self.exp is None:
            return
        if tab == 'scatter':
            scn = self.exp.scenarios[self.scenario_var.get()]
            opts = self._opts(tab)
            if self.exp.is_powercap and opts.get('mode') == 'cmp':
                fig = build_compare_figure(self.exp, scn, self.cmp_var.get(),
                                           self.styles, opts)
            else:
                fig = build_scatter_figure(self.exp, scn, self.styles, opts)
            self._mount_figure(tab, fig, pickable=True)
        elif tab == 'metrics':
            display = self.metric_var.get()
            if display in self._metric_defs:
                key, ylabel, source = self._metric_defs[display]
                fig = build_metric_figure(self.exp, key, display, ylabel, source,
                                          self.styles, self._opts(tab))
                self._mount_figure(tab, fig)
        elif tab == 'runtime':
            fig = build_runtime_figure(self.exp, self.styles, self._opts(tab))
            self._mount_figure(tab, fig)
        elif tab == 'feas':
            if self.exp.is_powercap:
                fig = build_feasibility_figure(self.exp, self.styles, self._opts(tab))
                self._mount_figure(tab, fig)
        elif tab == 'tables':
            self._refresh_table()

    def _mount_figure(self, tab, fig, pickable=False):
        frame = self.tabs[tab]
        old = self.canvases.get(tab)
        if old is not None:
            old.get_tk_widget().master.destroy()
        holder = self.ttk.Frame(frame)
        holder.pack(fill='both', expand=True)
        canvas = self.FigureCanvasTkAgg(fig, master=holder)
        canvas.draw()
        toolbar = self.NavigationToolbar2Tk(canvas, holder, pack_toolbar=False)
        toolbar.update()
        toolbar.pack(side='bottom', fill='x')
        canvas.get_tk_widget().pack(fill='both', expand=True)
        canvas.mpl_connect('scroll_event', self._on_scroll)
        if pickable:
            for artist in fig.axes[0].get_children():
                if artist.get_gid():
                    artist.set_picker(5)
            canvas.mpl_connect('pick_event', self._on_pick)
        self.figures[tab] = fig
        self.canvases[tab] = canvas

    def _on_scroll(self, event):
        ax = event.inaxes
        if ax is None:
            return
        factor = 1 / 1.25 if event.button == 'up' else 1.25
        for lim, get, set_, coord in (
                ('x', ax.get_xlim, ax.set_xlim, event.xdata),
                ('y', ax.get_ylim, ax.set_ylim, event.ydata)):
            lo, hi = get()
            if coord is None:
                coord = (lo + hi) / 2
            set_(coord - (coord - lo) * factor, coord + (hi - coord) * factor)
        event.canvas.draw_idle()

    # ---- scatter pick -> details -------------------------------------------

    def _powercap_full_label(self, base, tier):
        return base if tier == UNCAPPED_TIER else f'{base}_{tier}'

    def _on_pick(self, event):
        gid = event.artist.get_gid()
        if not gid or self.exp is None or not len(event.ind):
            return
        scn = self.exp.scenarios[self.scenario_var.get()]
        if not scn.details:
            self.set_status('No solution details in this run — Details tab unavailable.')
            return
        kind, name = gid.split('::', 1)
        idx = event.ind[0]
        powercap = self.exp.is_powercap
        cap_mode = powercap and self.mode_var.get() == 'cap'
        tier = self.tier_var.get() if cap_mode else None
        ox = scn.obj_names[1] if self.swap_var.get() else scn.obj_names[0]

        if kind == 'cloud':
            # In per-cap view the artist carries the BASE label and indexes the
            # tier's cloud rows; details are stored under the full tier label.
            pts_src = scn.tiers[tier].points if cap_mode else scn.points
            details_algo = self._powercap_full_label(name, tier) if cap_mode else name
            pts = pts_src[pts_src['Algorithm'] == name].reset_index()
            if idx >= len(pts):
                return
            row = pts.iloc[idx]
            seed = int(row['Seed'])
            sol_idx = int(row['index'] -
                          pts[pts['Seed'] == row['Seed']]['index'].min())
            self._select_details(details_algo, seed, sol_idx)
        elif kind == 'front':
            fronts_src = scn.tiers[tier].fronts if cap_mode else scn.fronts
            details_algo = self._powercap_full_label(name, tier) if cap_mode else name
            fr = fronts_src.get(name)
            if fr is None:
                return
            # Must mirror the builder's plot order (sorted by the CURRENT x axis).
            fr = fr.sort_values(ox).reset_index(drop=True)
            if idx >= len(fr):
                return
            target = fr.iloc[idx][scn.obj_names].to_numpy(dtype=float)
            self._select_details_by_objectives(details_algo, target, scn)
        elif kind == 'cmpfront':
            base = self.cmp_var.get()
            td = scn.tiers.get(name)
            fr = td.fronts.get(base) if td is not None else None
            if fr is None:
                return
            fr = fr.sort_values(ox).reset_index(drop=True)
            if idx >= len(fr):
                return
            target = fr.iloc[idx][scn.obj_names].to_numpy(dtype=float)
            self._select_details_by_objectives(
                self._powercap_full_label(base, name), target, scn)

    def _select_details_by_objectives(self, algo, target, scn):
        best, best_err = None, None
        for sol in details_solutions(scn):
            if sol.get('algorithm') != algo:
                continue
            objs = np.asarray(sol.get('objectives', []), dtype=float)
            if len(objs) != len(target):
                continue
            scale = np.maximum(np.abs(target), 1e-12)
            err = float(np.max(np.abs(objs - target) / scale))
            if best_err is None or err < best_err:
                best, best_err = sol, err
        if best is None or best_err > 1e-4:
            self.set_status(f'No matching captured solution for this {algo} point.')
            return
        self._select_details(algo, best['seed'], best['solutionIndex'])

    def _select_details(self, algo, seed, sol_idx):
        self.notebook.select(self.tabs['details'])
        self.d_algo_var.set(algo)
        self._details_fill_seeds()
        self.d_seed_var.set(str(seed))
        self._details_fill_solutions()
        for label in self.d_sol_combo['values']:
            if label.startswith(f'#{sol_idx} ') or label.startswith(f'#{sol_idx} '):
                self.d_sol_var.set(label)
                break
        else:
            values = self.d_sol_combo['values']
            if sol_idx < len(values):
                self.d_sol_var.set(values[sol_idx])
        self._details_show_selected()

    # ---- details tab ---------------------------------------------------------

    def _details_allowed_algorithms(self):
        """Full labels the Details pickers should offer in the current view
        (None = no filtering; powercap scopes to the tier / compared arm)."""
        exp = self.exp
        if exp is None or not exp.is_powercap:
            return None
        if self.mode_var.get() == 'cmp':
            base = self.cmp_var.get()
            return {self._powercap_full_label(base, t) for t in exp.tier_names}
        tier = self.tier_var.get()
        return {self._powercap_full_label(b, tier) for b in exp.base_algorithms}

    def _refresh_details_tab(self):
        scn = self.exp.scenarios[self.scenario_var.get()] if self.exp else None
        if scn is None or not scn.details:
            self.details_body.pack_forget()
            self.details_msg.pack(fill='both', expand=True)
            return
        self.details_msg.pack_forget()
        self.details_body.pack(fill='both', expand=True)
        allowed = self._details_allowed_algorithms()
        algos = sorted({s.get('algorithm') for s in details_solutions(scn)
                        if allowed is None or s.get('algorithm') in allowed})
        self.d_algo_combo['values'] = algos
        if self.d_algo_var.get() not in algos and algos:
            self.d_algo_var.set(algos[0])
        self._details_fill_seeds()
        self._details_fill_solutions()
        self._details_show_selected()

    def _details_filter_changed(self):
        self._details_fill_seeds()
        self._details_fill_solutions()
        self._details_show_selected()

    def _details_fill_seeds(self):
        scn = self.exp.scenarios[self.scenario_var.get()]
        seeds = sorted({s.get('seed') for s in details_solutions(scn)
                        if s.get('algorithm') == self.d_algo_var.get()})
        self.d_seed_combo['values'] = [str(s) for s in seeds]
        if self.d_seed_var.get() not in self.d_seed_combo['values'] and seeds:
            self.d_seed_var.set(str(seeds[0]))

    def _details_fill_solutions(self):
        scn = self.exp.scenarios[self.scenario_var.get()]
        obj_names = (scn.details or {}).get('objectives', scn.obj_names)
        try:
            seed = int(self.d_seed_var.get())
        except ValueError:
            seed = None
        self._details_sols = [
            s for s in details_solutions(scn)
            if s.get('algorithm') == self.d_algo_var.get() and s.get('seed') == seed]
        self._details_sols.sort(key=lambda s: s.get('solutionIndex', 0))
        labels = [solution_label(s, obj_names) for s in self._details_sols]
        self.d_sol_combo['values'] = labels
        if labels and self.d_sol_var.get() not in labels:
            self.d_sol_var.set(labels[0])

    def _details_show_selected(self):
        if not self._details_sols:
            for tree in (self.d_summary_tree, self.d_vm_tree,
                         self.d_host_tree, self.d_task_tree):
                tree.delete(*tree.get_children())
            return
        try:
            idx = list(self.d_sol_combo['values']).index(self.d_sol_var.get())
        except ValueError:
            idx = 0
        sol = self._details_sols[idx]
        scn = self.exp.scenarios[self.scenario_var.get()]

        for tree, rows in (
                (self.d_summary_tree, details_summary_rows(sol, scn)),
                (self.d_vm_tree, details_vm_table(sol, scn)),
                (self.d_host_tree, details_host_table(sol, scn)),
                (self.d_task_tree, details_task_table(sol, scn))):
            tree.delete(*tree.get_children())
            for row in rows:
                tree.insert('', 'end', values=row)
        total = sum(sol.get('hostEnergyKWh') or [])
        self.d_host_tree.insert('', 'end',
                                values=('TOTAL', '', '', '', f'{total:.9f}'))
        self.set_status(
            f'Details: {sol.get("algorithm")} seed {sol.get("seed")} '
            f'solution #{sol.get("solutionIndex")}')

    # ---- tables tab ------------------------------------------------------------

    # Column-aware cell formatting for the Tables tab / exports.
    @staticmethod
    def _format_cell(col, v):
        if pd.isna(v):
            return ''
        if col in ('Front size', 'Contribution', 'Runtime (ms)'):
            return f'{v:.0f}'
        if col == '% of universal':
            return f'{v:.1f}'
        if col == 'Avg runtime (s)':
            return f'{v:.2f}'
        return f'{v:.6f}'

    def _table_source(self, scn):
        """ScenarioData, or the selected TierData in powercap per-cap view."""
        if self.exp.is_powercap and self.mode_var.get() != 'cmp':
            return scn.tiers.get(self.tier_var.get(), scn)
        return scn

    def _refresh_table(self):
        scn = self.exp.scenarios[self.scenario_var.get()]
        kind = self.table_kind_var.get()
        if self.exp.is_powercap and self.mode_var.get() == 'cmp':
            metric = COMPARE_TABLE_KINDS.get(kind)
            df, info = (compare_pivot_table(self.exp, scn, self.cmp_var.get(), metric)
                        if metric else (None, ''))
        else:
            # TierData quacks like ScenarioData for the table builders, so the
            # per-cap view just swaps the source object.
            src = self._table_source(scn)
            if kind == GLOBAL_TABLE_KIND:
                algos = (self.exp.base_algorithms if self.exp.is_powercap
                         else self.exp.algorithms)
                df, info = global_metric_table(src, scn.obj_names, algos)
            elif kind.startswith('Per-seed — '):
                metric = PER_SEED_TABLE_COLS.get(kind.split('— ', 1)[1])
                df, info = (pivot_metric_table(src, metric) if metric
                            else (None, ''))
            else:
                df, info = None, ''
            if (df is not None and self.ref_var.get()
                    and kind in (GLOBAL_TABLE_KIND, 'Per-seed — Hypervolume')):
                info = (info + '\n' if info else '') + hv_frame_info(
                    _src_union_bounds(src, scn.obj_names), scn.obj_names)
        self.current_table = (df, info)
        tree = self.table_tree
        tree.delete(*tree.get_children())
        if df is None:
            self.table_info.config(text='Not available in this run.')
            return
        by_algorithm = df.index.name == 'Algorithm'  # Global metrics table
        count_table = kind.split('— ', 1)[-1] in ('Front size',
                                                  'Pareto contribution',
                                                  'Runtime (ms)')
        first_col = df.index.name or 'Seed'
        cols = [first_col] + [self.styles.get(c, {'display': c})['display']
                              for c in df.columns]
        tree['columns'] = cols
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=max(90, 9 * len(col)), anchor='e')
        tree.column(first_col, width=200 if by_algorithm else 70, anchor='w')
        for idx, row in df.iterrows():
            agg = idx in ('MEAN', 'STDDEV', UNIVERSAL_KEY)
            name = (UNIVERSAL_LABEL if idx == UNIVERSAL_KEY else
                    self.styles.get(idx, {}).get('display', idx)
                    if by_algorithm else idx)
            if by_algorithm:
                cells = [self._format_cell(c, v) for c, v in zip(df.columns, row)]
            elif count_table:
                cells = ['' if pd.isna(v) else f'{v:g}' for v in row]
            else:
                cells = ['' if pd.isna(v) else f'{v:.6f}' for v in row]
            tree.insert('', 'end', values=[name] + cells,
                        tags=('agg',) if agg else ())
        self.table_info.config(text=info or '')

    # ---- saving -----------------------------------------------------------------

    def save_figure(self):
        tab = self._current_tab()
        if tab not in PLOT_TABS or tab not in self.figures:
            self.messagebox.showinfo(
                'Save figure', 'Switch to a plot tab (Scatter / Metrics / Runtime) first.')
            return
        path = self.filedialog.asksaveasfilename(
            defaultextension='.png',
            initialfile=f'{tab}_{self.exp.name}.png',
            filetypes=[('PNG', '*.png'), ('SVG', '*.svg'), ('PDF', '*.pdf')])
        if not path:
            return
        try:
            dpi = float(self.dpi_var.get())
        except ValueError:
            dpi = 300
        self.figures[tab].savefig(path, dpi=dpi, bbox_inches='tight')
        self.set_status(f'Saved {path}')

    def save_table(self):
        tab = self._current_tab()
        if tab == 'details' and self._details_sols:
            self._save_details_tables()
            return
        if self.current_table is None or self.current_table[0] is None:
            self.messagebox.showinfo('Save table', 'Open the Tables tab first.')
            return
        df, info = self.current_table
        path = self.filedialog.asksaveasfilename(
            defaultextension='.csv', initialfile=f'table_{self.exp.name}.csv',
            filetypes=[('CSV', '*.csv'), ('Excel workbook', '*.xlsx')])
        if not path:
            return
        if path.lower().endswith('.xlsx'):
            if not self._check_openpyxl():
                return
            write_single_table_xlsx(df, info, path,
                                    display_map=self._display_map())
        else:
            with open(path, 'w') as fh:
                df.to_csv(fh, index_label=df.index.name or 'Seed',
                          float_format='%.6f')
                if info:
                    fh.write(f'\n# {info}\n')
        self.set_status(f'Saved {path}')

    def _save_details_tables(self):
        outdir = self.filedialog.askdirectory(
            title='Directory for the four detail CSVs of the selected solution')
        if not outdir:
            return
        try:
            idx = list(self.d_sol_combo['values']).index(self.d_sol_var.get())
        except ValueError:
            idx = 0
        sol = self._details_sols[idx]
        scn = self.exp.scenarios[self.scenario_var.get()]
        stem = (f'{sol.get("algorithm")}_seed{sol.get("seed")}'
                f'_sol{sol.get("solutionIndex")}')
        pd.DataFrame(details_summary_rows(sol, scn),
                     columns=['Metric', 'Value']).to_csv(
            os.path.join(outdir, f'{stem}_summary.csv'), index=False)
        pd.DataFrame(details_vm_table(sol, scn),
                     columns=['VM', 'Type', 'vCPUs', 'GPUs', 'Host',
                              'QueueSize', 'QueueOrder']).to_csv(
            os.path.join(outdir, f'{stem}_vm_queues.csv'), index=False)
        pd.DataFrame(details_host_table(sol, scn),
                     columns=['Host', 'Type', 'Cores', 'GPUs',
                              'EnergyKWh']).to_csv(
            os.path.join(outdir, f'{stem}_host_energy.csv'), index=False)
        pd.DataFrame(details_task_table(sol, scn),
                     columns=['Task', 'LengthInstructions', 'Workload', 'VM',
                              'WaitSeconds']).to_csv(
            os.path.join(outdir, f'{stem}_task_assignments.csv'), index=False)
        self.set_status(f'Saved 4 detail CSVs to {outdir}')

    # ---- Excel export -----------------------------------------------------

    def _display_map(self):
        return {a: st['display'] for a, st in self.styles.items()}

    def _check_openpyxl(self):
        if have_openpyxl():
            return True
        self.messagebox.showerror('Excel export', OPENPYXL_HINT)
        return False

    def export_table_xlsx(self):
        """The table currently shown in the Tables tab -> one-sheet .xlsx."""
        if self.exp is None:
            return
        if self.current_table is None or self.current_table[0] is None:
            self.messagebox.showinfo('Excel export', 'Open the Tables tab first.')
            return
        if not self._check_openpyxl():
            return
        kind = self.table_kind_var.get()
        sheet = ('Global' if kind == GLOBAL_TABLE_KIND
                 else kind.split('— ', 1)[-1])
        path = self.filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            initialfile=f'table_{self.exp.name}.xlsx',
            filetypes=[('Excel workbook', '*.xlsx')])
        if not path:
            return
        df, info = self.current_table
        write_single_table_xlsx(df, info, path,
                                display_map=self._display_map(), sheet=sheet)
        self.set_status(f'Saved {path}')

    def export_workbook_dialog(self):
        """Every table of every scenario (and cap tier) -> one workbook."""
        if self.exp is None:
            return
        if not self._check_openpyxl():
            return
        path = self.filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            initialfile=f'{self.exp.name}_tables.xlsx',
            filetypes=[('Excel workbook', '*.xlsx')])
        if not path:
            return
        n = export_tables_xlsx(self.exp, path)
        self.set_status(f'Saved {n} table sheets + README to {path}')

    def export_all_dialog(self):
        if self.exp is None:
            return
        outdir = self.filedialog.askdirectory(title='Export all figures + tables to…')
        if not outdir:
            return
        try:
            dpi = float(self.dpi_var.get())
        except ValueError:
            dpi = 300
        written = export_all(self.exp, outdir, dpi=dpi)
        self.set_status(f'Exported {len(written)} files to {outdir}')

    def save_for_claude(self):
        if self.exp is None:
            return
        path = self.filedialog.asksaveasfilename(
            defaultextension='.zip',
            initialfile=f'claude_bundle_{self.exp.name}.zip',
            filetypes=[('Claude bundle (zip: briefing + figures + tables)', '*.zip'),
                       ('Markdown only', '*.md'),
                       ('Gzipped Markdown only', '*.md.gz')])
        if not path:
            return
        if path.endswith('.zip'):
            try:
                dpi = float(self.dpi_var.get())
            except ValueError:
                dpi = 300
            # Bundles default to 200 DPI (documented) unless the spinner was
            # moved off its default.
            if dpi == 300:
                dpi = 200
            self.set_status('Building Claude bundle zip (renders every figure)…')
            self.root.update_idletasks()
            export_claude_zip(self.exp, path, dpi=dpi)
        else:
            text = build_claude_bundle(self.exp)
            if path.endswith('.gz'):
                with gzip.open(path, 'wt') as fh:
                    fh.write(text)
            else:
                with open(path, 'w') as fh:
                    fh.write(text)
        self.set_status(f'Saved Claude bundle: {path} '
                        f'({os.path.getsize(path) / 1024:.0f} KB)')


def launch_gui(folder=None):
    try:
        import tkinter as tk
        from tkinter import ttk, filedialog, colorchooser, messagebox
    except ImportError:
        print('tkinter is not available. Install python3-tk, or use '
              '--export-all <dir> for headless export.', file=sys.stderr)
        return 2
    from matplotlib.backends.backend_tkagg import (
        FigureCanvasTkAgg, NavigationToolbar2Tk)
    app = ResultsExplorerApp({
        'tk': tk, 'ttk': ttk, 'filedialog': filedialog,
        'colorchooser': colorchooser, 'messagebox': messagebox,
        'FigureCanvasTkAgg': FigureCanvasTkAgg,
        'NavigationToolbar2Tk': NavigationToolbar2Tk,
    }, folder=folder)
    app.run()
    return 0


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Interactive explorer for campaign experiment results '
                    '(manual tool; never auto-run by the pipeline).')
    parser.add_argument('folder', nargs='?',
                        help='experiment results folder (results/<ExperimentId>)')
    parser.add_argument('--export-all', metavar='DIR',
                        help='headless: export all figures/tables/bundle to DIR and exit')
    parser.add_argument('--export-xlsx', metavar='FILE.xlsx',
                        help='headless: write every table into one formatted '
                             'Excel workbook (needs openpyxl) and exit')
    parser.add_argument('--claude-zip', metavar='FILE.zip',
                        help='headless: write the self-briefing Claude bundle zip '
                             '(instructions + bundle.md + figures/ + tables/) and exit')
    parser.add_argument('--dpi', type=float, default=300.0,
                        help='DPI for --export-all figures (default 300; '
                             'the Claude zip uses 200 unless overridden)')
    args = parser.parse_args(argv)

    if args.export_all or args.claude_zip or args.export_xlsx:
        if not args.folder:
            parser.error('--export-all/--export-xlsx/--claude-zip require '
                         'a results folder argument')
        exp = ExperimentData.load(args.folder)
        if args.export_all:
            written = export_all(exp, args.export_all, dpi=args.dpi)
            print(f'Exported {len(written)} files to {args.export_all}')
        if args.export_xlsx:
            try:
                n = export_tables_xlsx(exp, args.export_xlsx)
            except ImportError:
                print(OPENPYXL_HINT, file=sys.stderr)
                return 2
            print(f'Wrote {n} table sheets + README to {args.export_xlsx}')
        if args.claude_zip:
            dpi = args.dpi if args.dpi != 300.0 else 200.0
            export_claude_zip(exp, args.claude_zip, dpi=dpi)
            size = os.path.getsize(args.claude_zip) / (1024 * 1024)
            print(f'Wrote Claude bundle: {args.claude_zip} ({size:.1f} MB)')
        return 0
    return launch_gui(args.folder)


if __name__ == '__main__':
    sys.exit(main())

