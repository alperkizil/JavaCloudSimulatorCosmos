#!/usr/bin/env python3
"""
Regression test for results_explorer.py's simplified single-HV UI surface:
the Global metrics (collective pooled-front) table, the in-explorer HV_fixed
fallback for folders that pre-date the CSV column, the collective/mean
aggregation of the metric figures, the HV reference-point overlay, and the
Excel export.

Usage:
    python scripts/test_results_explorer_global_tables.py <plain_folder> [powercap_folder]

<plain_folder> should be an ordinary campaign folder (e.g. one of the
newExperimentResults reference runs — with or without a native HV_fixed
column). [powercap_folder] optionally exercises the per-tier surface (e.g. a
folder written by SyntheticPowerCeilingFolder).

Exit code 0 = all checks passed. Needs matplotlib (Agg is forced); the Excel
checks are skipped with a notice when openpyxl is not installed.
"""

import os
import sys
import tempfile

import matplotlib
matplotlib.use('Agg')
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from results_explorer import (  # noqa: E402
    ExperimentData, GLOBAL_COLUMNS, GLOBAL_TABLE_KIND, METRIC_CHOICES,
    PER_SEED_TABLES, TABLE_METRICS, UNIVERSAL_KEY, available_metrics,
    build_metric_figure, build_scatter_figure, collective_metrics, export_all,
    global_metric_table, have_openpyxl, hv_frame_info, pivot_metric_table,
    powercap_metrics, resolve_styles, _src_union_bounds)

FAILURES = 0
CHECKS = 0


def check(what, ok, detail=''):
    global FAILURES, CHECKS
    CHECKS += 1
    if not ok:
        FAILURES += 1
        print(f'  FAIL {what}{": " + detail if detail else ""}')


def check_single_hv_surface():
    """The choice lists offer exactly one hypervolume and one contribution."""
    keys = [k for k, _, _ in METRIC_CHOICES]
    check('metric keys are the simple five',
          keys == ['HV_fixed', 'GD', 'IGD', 'Spacing', 'ParetoContribution'],
          str(keys))
    check('no legacy HV / TotalSolutions / collab in per-seed tables',
          [c for _, c in PER_SEED_TABLES] ==
          ['HV_fixed', 'GD', 'IGD', 'Spacing', 'NonDomSolutions',
           'ParetoContribution', 'TimeMs'])
    check('bulk-export pivots carry no legacy HV', 'HV' not in TABLE_METRICS
          and 'TotalSolutions' not in TABLE_METRICS)


def check_global_table(src, obj_names, algorithms, tag, metrics_mean):
    df, info = global_metric_table(src, obj_names, algorithms)
    check(f'{tag} global table built', df is not None)
    if df is None:
        return
    check(f'{tag} global columns', list(df.columns) == GLOBAL_COLUMNS,
          str(list(df.columns)))
    check(f'{tag} universal row present', UNIVERSAL_KEY in df.index)
    uni = df.loc[UNIVERSAL_KEY]
    check(f'{tag} universal row self-consistent',
          uni['Contribution'] == uni['Front size']
          and uni['% of universal'] == 100.0
          and uni['GD'] == 0.0 and uni['IGD'] == 0.0)
    check(f'{tag} universal Hypervolume is the ceiling',
          df['Hypervolume'].idxmax() == UNIVERSAL_KEY)
    for algo in algorithms:
        if algo not in df.index:
            continue
        # Collective front size equals the pooled front's ND size.
        fr = src.fronts.get(algo)
        if fr is not None:
            check(f'{tag} {algo} front size vs pooled front',
                  df.loc[algo, 'Front size'] <= len(fr) and
                  df.loc[algo, 'Front size'] > 0)
        # Strict contribution equals the CSV MEAN row's union count.
        if metrics_mean is not None and algo in metrics_mean.index:
            check(f'{tag} {algo} contribution == MEAN-row union count',
                  int(df.loc[algo, 'Contribution'])
                  == int(metrics_mean.loc[algo, 'ParetoContribution']))
        # A union front can only match or beat every one of its seed fronts.
        if (src.metrics_seed is not None
                and 'HV_fixed' in src.metrics_seed.columns):
            per_seed = src.metrics_seed.loc[
                src.metrics_seed['Algorithm'] == algo, 'HV_fixed'].dropna()
            if len(per_seed):
                check(f'{tag} {algo} collective HV >= best per-seed HV',
                      df.loc[algo, 'Hypervolume'] >= per_seed.max() - 1e-9,
                      f'{df.loc[algo, "Hypervolume"]:.6f} < {per_seed.max():.6f}')


def check_figures(exp, tag):
    styles = resolve_styles(exp.base_algorithms if exp.is_powercap
                            else exp.algorithms)
    choices = powercap_metrics(exp) if exp.is_powercap else available_metrics(exp)
    opts_extra = {'tier': exp.tier_names[0], 'mode': 'cap'} if exp.is_powercap else {}
    for key, display, ylabel in choices:
        for agg in ('collective', 'mean'):
            fig = build_metric_figure(exp, key, display, ylabel, None, styles,
                                      dict(opts_extra, aggregation=agg))
            check(f'{tag} {key}/{agg} figure', len(fig.axes) >= 1)
    scn = next(iter(exp.scenarios.values()))
    for norm in (False, True):
        fig = build_scatter_figure(exp, scn, styles,
                                   dict(opts_extra, show_ref_point=True,
                                        normalize=norm))
        legend = fig.axes[0].get_legend()
        labels = [t.get_text() for t in legend.get_texts()] if legend else []
        check(f'{tag} scatter ref point (normalize={norm})',
              'HV reference point' in labels, str(labels))


def check_hv_fallback(exp, tag):
    """When the folder has no HV_fixed column the explorer must fill it in;
    verify against an independent implementation of the recompute_hv math."""
    for n, scn in exp.scenarios.items():
        check(f'{tag} s{n} has HV_fixed', scn.has_hv_fixed
              and 'HV_fixed' in scn.metrics_seed.columns)
        if not exp.hv_fixed_recomputed:
            continue
        bounds = _src_union_bounds(scn, scn.obj_names)
        ix, iy, nx, ny = bounds
        rx = max(nx - ix, 1e-12)
        ry = max(ny - iy, 1e-12)
        # Independent HV: staircase sweep against (1.1, 1.1) normalized.
        for algo, grp in list(scn.points.groupby('Algorithm', sort=False))[:2]:
            for seed, sub in list(grp.groupby('Seed', sort=False))[:2]:
                pts = sub[scn.obj_names].to_numpy(dtype=float)
                norm = np.column_stack(((pts[:, 0] - ix) / rx,
                                        (pts[:, 1] - iy) / ry))
                keep = [p for i, p in enumerate(norm)
                        if not any((q[0] <= p[0] and q[1] <= p[1]
                                    and (q[0] < p[0] or q[1] < p[1]))
                                   for j, q in enumerate(norm) if i != j)]
                nd = np.unique(np.round(np.array(keep), 9), axis=0)
                nd = nd[(nd[:, 0] < 1.1) & (nd[:, 1] < 1.1)]
                nd = nd[np.argsort(nd[:, 0])]
                hv, prev_y = 0.0, 1.1
                for x, y in nd:
                    if y < prev_y:
                        hv += (1.1 - x) * (prev_y - y)
                        prev_y = y
                expected = hv / 1.21
                got = scn.metrics_seed.loc[
                    (scn.metrics_seed['Algorithm'] == algo)
                    & (scn.metrics_seed['Seed'] == int(seed)), 'HV_fixed'].iloc[0]
                check(f'{tag} s{n} {algo}/{seed} fallback HV value',
                      abs(got - expected) < 1e-9,
                      f'got {got:.9f} expected {expected:.9f}')
        # MEAN/STDDEV rows and the universal trailer were filled in too.
        check(f'{tag} s{n} MEAN row got HV_fixed',
              scn.metrics_mean is not None
              and scn.metrics_mean['HV_fixed'].notna().all())
        check(f'{tag} s{n} universal trailer got HV_fixed',
              pd_notna(scn.universal_metrics.get('HV_fixed')))


def pd_notna(v):
    return v is not None and not (isinstance(v, float) and np.isnan(v))


def check_exports(exp, tag):
    with tempfile.TemporaryDirectory() as tmp:
        written = export_all(exp, tmp, dpi=60)
        names = [os.path.basename(p) for p in written]
        check(f'{tag} export-all has global CSVs',
              any('global_metrics' in n for n in names))
        check(f'{tag} export-all has both aggregations',
              any(n.endswith('_collective.png') for n in names)
              and any(n.endswith('_mean.png') for n in names))
        check(f'{tag} export-all has no collab / legacy-HV tables',
              not any('seed_collaboration' in n for n in names)
              and not any(n.endswith('_HV.csv')
                          or '_HV_' in n.replace('_HV_fixed', '')
                          for n in names if n.startswith('table_')))
        if not have_openpyxl():
            print(f'  note: openpyxl not installed — {tag} Excel checks skipped')
            return
        from results_explorer import export_tables_xlsx
        import openpyxl
        path = os.path.join(tmp, 'tables.xlsx')
        n_sheets = export_tables_xlsx(exp, path)
        wb = openpyxl.load_workbook(path)
        check(f'{tag} workbook sheet count', len(wb.sheetnames) == n_sheets + 1,
              f'{len(wb.sheetnames)} vs {n_sheets}+README')
        check(f'{tag} workbook README first', wb.sheetnames[0] == 'README')
        globals_ = [s for s in wb.sheetnames if s.endswith('Global')]
        check(f'{tag} workbook has Global sheets', len(globals_) > 0)
        ws = wb[globals_[0]]
        check(f'{tag} sheet frozen header', ws.freeze_panes == 'B2')
        shaded = [row[0].value for row in ws.iter_rows()
                  if row[0].fill.start_color.rgb
                  and str(row[0].fill.start_color.rgb).endswith('DDE8F0')]
        check(f'{tag} universal row shaded', 'Universal Pareto' in shaded,
              str(shaded))


def run_plain(folder):
    print(f'-- Plain folder: {folder}')
    exp = ExperimentData.load(folder)
    scn = next(iter(exp.scenarios.values()))
    check_hv_fallback(exp, 'plain')
    for n, s in exp.scenarios.items():
        check_global_table(s, s.obj_names, exp.algorithms, f'plain s{n}',
                           s.metrics_mean)
    check_figures(exp, 'plain')
    check_exports(exp, 'plain')
    # Info surfaces.
    piv, info = pivot_metric_table(scn, 'HV_fixed')
    check('plain per-seed pivot info uses Hypervolume wording',
          piv is not None and 'Hypervolume=' in info, info)
    check('plain frame info mentions the reference point',
          'reference point' in hv_frame_info(
              _src_union_bounds(scn, scn.obj_names), scn.obj_names))


def run_powercap(folder):
    print(f'-- PowerCap folder: {folder}')
    exp = ExperimentData.load(folder)
    check('powercap detected', exp.is_powercap)
    for n, scn in exp.scenarios.items():
        for tier in exp.tier_names:
            td = scn.tiers.get(tier)
            if td is None:
                continue
            check_global_table(td, scn.obj_names, exp.base_algorithms,
                               f'pc s{n}/{tier}', td.metrics_mean)
            rows, uni_row, _ = collective_metrics(td, scn.obj_names,
                                                  exp.base_algorithms)
            check(f'pc s{n}/{tier} tier universal sized like TierData',
                  uni_row.get('Front size') == len(td.universal))
    check_figures(exp, 'pc')
    check_exports(exp, 'pc')


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    check_single_hv_surface()
    run_plain(sys.argv[1])
    if len(sys.argv) > 2:
        run_powercap(sys.argv[2])
    print()
    print(f'Checks: {CHECKS}, failures: {FAILURES}')
    print('=== PASSED ===' if FAILURES == 0 else '=== FAILED ===')
    return 0 if FAILURES == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
